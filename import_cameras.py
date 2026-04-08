import os
import openpyxl
import re

from app.db.session import SessionLocal
from app.db.models import Asset

directory_path = r"c:\Users\Administrator\Desktop\camera checklist"

def import_cameras():
    db = SessionLocal()
    total_imported = 0
    
    for root, _, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~'):
                filepath = os.path.join(root, file)
                print(f"Reading {filepath}...")
                try:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    ws = wb.active
                    
                    nvr_ip = None
                    # Fallback to filename as IP if NVR string is missing
                    filename_ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', file)
                    if filename_ip_match:
                        nvr_ip = filename_ip_match.group(1)
                    
                    camera_names = []
                    
                    # Try to extract exact NVR IP from row 4
                    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
                    if row4[0] and 'ĐẦU GHI' in str(row4[0]):
                        ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', str(row4[0]))
                        if ip_match:
                            nvr_ip = ip_match.group(1)
                            
                    if not nvr_ip:
                         nvr_ip = file.replace('.xlsx', '')
                    
                    folder_name = os.path.basename(os.path.dirname(filepath))
                    base_filename = os.path.splitext(file)[0]
                    # Expected format: "Folder_FileName"
                    expected_nvr_name = f"{folder_name}_{base_filename}"
                    
                    # Ensure NVR asset exists
                    nvr_asset = db.query(Asset).filter(Asset.asset_type == 'NVR', Asset.ip_address == nvr_ip).first()
                    if not nvr_asset:
                        nvr_asset = Asset(
                            asset_code=f"NVR-{nvr_ip.replace('.', '')}",
                            asset_name=expected_nvr_name,
                            asset_type="NVR",
                            ip_address=nvr_ip,
                            status="active"
                        )
                        db.add(nvr_asset)
                        db.flush()
                        print(f"  Created NVR: {nvr_ip} as {expected_nvr_name}")
                    else:
                        if nvr_asset.asset_name != expected_nvr_name:
                            print(f"  Updating NVR name from '{nvr_asset.asset_name}' to '{expected_nvr_name}'")
                            nvr_asset.asset_name = expected_nvr_name


                    # Extract Camera Names from row 9
                    row9 = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]
                    # Columns start from index 2 (D1)
                    for i in range(2, len(row9)):
                        cam_name = row9[i]
                        if cam_name and isinstance(cam_name, str) and cam_name.strip():
                            # Clean name
                            clean_name = cam_name.strip()
                            idx = i - 1 # D1 = 1
                            camera_names.append((idx, clean_name))
                            
                    for idx, name in camera_names:
                        cam_code = f"CAM-{nvr_ip.replace('.', '')}-D{idx}"
                        existing_cam = db.query(Asset).filter(Asset.asset_code == cam_code).first()
                        if not existing_cam:
                            cam_asset = Asset(
                                asset_code=cam_code,
                                asset_name=name,
                                asset_type="Camera",
                                location=nvr_ip,
                                status="active"
                            )
                            db.add(cam_asset)
                            total_imported += 1
                            print(f"  + Added Camera {idx}: {name}")
                except Exception as e:
                    print(f"Error processing {file}: {e}")

    db.commit()
    db.close()
    print(f"Successfully imported {total_imported} cameras.")

if __name__ == '__main__':
    import_cameras()
