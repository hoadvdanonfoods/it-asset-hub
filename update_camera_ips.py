"""
Script cập nhật IP Camera từ file Excel checklist vào database.
Quét dòng "Ip Tương Ứng" trong mỗi file Excel và ghép với camera tương ứng.
"""
import os
import re
import openpyxl
from app.db.session import SessionLocal
from app.db.models import Asset

CHECKLIST_DIR = "data/camera_checklists"


def find_ip_row(ws, max_scan=15):
    """Tìm dòng chứa 'Ip Tương Ứng' hoặc 'IP' trong cột A/B."""
    for r in range(1, max_scan + 1):
        for c in range(1, 4):
            val = ws.cell(row=r, column=c).value
            if val and ('Ip Tương Ứng' in str(val) or 'IP Tương Ứng' in str(val).upper()
                        or str(val).strip().lower() in ('ip tương ứng', 'ip tuong ung', 'ip camera')):
                return r
    return None


def find_d_header_row(ws, max_scan=15):
    """Tìm dòng chứa D1, D2,... (header mã camera)."""
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip() == 'D1':
                return r
    return None


def update_camera_ips():
    db = SessionLocal()
    updated = 0
    skipped = 0

    for root, dirs, files in os.walk(CHECKLIST_DIR):
        for file in files:
            if not file.endswith('.xlsx') or file.startswith('~'):
                continue

            filepath = os.path.join(root, file)
            print(f"\n📂 Reading: {filepath}")

            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
            except Exception as load_err:
                print(f"  ⚠ Lỗi mở file: {load_err}")
                continue

            try:
                ws = wb.active

                ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', file)
                if not ip_match:
                    # Thử tìm trong row 4
                    row4_val = ws.cell(row=4, column=1).value
                    if row4_val:
                        ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', str(row4_val))
                if not ip_match:
                    print(f"  ⚠ Không tìm thấy NVR IP trong file")
                    wb.close()
                    continue

                nvr_ip = ip_match.group(1)
                print(f"  NVR IP: {nvr_ip}")

                # Tìm dòng D-header (D1, D2,...)
                d_row = find_d_header_row(ws)
                if not d_row:
                    print(f"  ⚠ Không tìm thấy dòng D1/D2 header")
                    wb.close()
                    continue
                print(f"  D-header row: {d_row}")

                # Tìm dòng IP tương ứng
                ip_row = find_ip_row(ws)
                if not ip_row:
                    print(f"  ⚠ Không tìm thấy dòng 'Ip Tương Ứng'")
                    wb.close()
                    continue
                print(f"  IP row: {ip_row}")

                # Xây dựng map D-code -> column
                d_col_map = {}
                for c in range(1, ws.max_column + 1):
                    hdr = ws.cell(row=d_row, column=c).value
                    if hdr and str(hdr).strip().startswith('D'):
                        d_col_map[str(hdr).strip()] = c

                # Đọc IP từng camera
                for d_code, col in d_col_map.items():
                    cam_ip_raw = ws.cell(row=ip_row, column=col).value
                    if not cam_ip_raw:
                        continue

                    cam_ip = str(cam_ip_raw).strip()
                    # Nếu IP không đầy đủ (VD: .40.1), ghép với prefix của NVR IP
                    if cam_ip.startswith('.'):
                        # Lấy prefix từ NVR IP: 192.168 từ 192.168.40.250
                        parts = nvr_ip.split('.')
                        # Đếm số phần trong cam_ip
                        ip_parts = [p for p in cam_ip.split('.') if p]
                        missing_parts = 4 - len(ip_parts)
                        prefix = '.'.join(parts[:missing_parts])
                        cam_ip = prefix + cam_ip

                    # Tìm camera trong DB theo asset_code
                    d_idx = d_code.replace('D', '')
                    cam_code = f"CAM-{nvr_ip.replace('.', '')}-D{d_idx}"
                    cam = db.query(Asset).filter(Asset.asset_code == cam_code).first()

                    if cam:
                        if cam.ip_address != cam_ip:
                            print(f"  ✅ {cam_code} ({cam.asset_name}): IP = {cam_ip}")
                            cam.ip_address = cam_ip
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        print(f"  ❌ Camera {cam_code} không tìm thấy trong DB")

                wb.close()
            except Exception as e:
                print(f"  ❗ Lỗi: {e}")

    db.commit()
    db.close()
    print(f"\n{'='*50}")
    print(f"✅ Đã cập nhật IP cho {updated} camera")
    print(f"⏭ Bỏ qua {skipped} camera (IP đã đúng)")


if __name__ == '__main__':
    update_camera_ips()
