#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import re
import unicodedata
from collections import OrderedDict

import openpyxl
from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
if REPO_ROOT.name != 'it-asset-hub':
    REPO_ROOT = REPO_ROOT / 'it-asset-hub'

SOURCE_XLSX = Path('/mnt/c/Users/Administrator/Desktop/DS nhân viên.xlsx')
STATE_DIR = REPO_ROOT / 'state'
EMP_OUTPUT = STATE_DIR / 'employees_import_ready.xlsx'
DEPT_OUTPUT = STATE_DIR / 'departments_import_ready.xlsx'
EMP_DESKTOP = Path('/mnt/c/Users/Administrator/Desktop/employees_import_ready.xlsx')
DEPT_DESKTOP = Path('/mnt/c/Users/Administrator/Desktop/departments_import_ready.xlsx')


def slugify(value: str) -> str:
    text = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^A-Za-z0-9]+', '-', text).strip('-')
    text = re.sub(r'-+', '-', text)
    return text.upper() or 'DEPT'


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def main() -> int:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.active

    headers = [clean_text(c.value) for c in ws[1]]
    idx_map = {name: i for i, name in enumerate(headers)}

    employees = []
    departments = OrderedDict()

    for row in ws.iter_rows(min_row=2, values_only=True):
        employee_code = clean_text(row[idx_map['Mã NV']])
        full_name = clean_text(row[idx_map['Họ và tên']])
        department_name = clean_text(row[idx_map['Tên phòng ban']])
        if not employee_code or not full_name:
            continue
        if department_name and department_name not in departments:
            departments[department_name] = {
                'code': slugify(department_name)[:20],
                'name': department_name,
                'is_active': True,
                'note': 'Chuẩn hóa từ DS nhân viên.xlsx',
            }
        employees.append({
            'employee_code': employee_code,
            'full_name': full_name,
            'department_name': department_name,
            'title': None,
            'email': None,
            'phone': None,
            'is_active': True,
            'note': 'Chuẩn hóa từ DS nhân viên.xlsx',
        })

    dept_id_map = {name: idx for idx, name in enumerate(departments.keys(), start=1)}

    dept_wb = Workbook()
    dept_ws = dept_wb.active
    dept_ws.title = 'Departments'
    dept_ws.append(['code', 'name', 'is_active', 'note'])
    for department in departments.values():
        dept_ws.append([department['code'], department['name'], department['is_active'], department['note']])

    emp_wb = Workbook()
    emp_ws = emp_wb.active
    emp_ws.title = 'Employees'
    emp_ws.append(['employee_code', 'full_name', 'department_id', 'title', 'email', 'phone', 'is_active', 'note'])
    for employee in employees:
        emp_ws.append([
            employee['employee_code'],
            employee['full_name'],
            dept_id_map.get(employee['department_name']),
            employee['title'],
            employee['email'],
            employee['phone'],
            employee['is_active'],
            employee['note'],
        ])

    review_ws = emp_wb.create_sheet('DepartmentMap')
    review_ws.append(['department_id', 'department_name', 'department_code'])
    for name, department in departments.items():
        review_ws.append([dept_id_map[name], name, department['code']])

    STATE_DIR.mkdir(parents=True, exist_ok=True)
    dept_wb.save(DEPT_OUTPUT)
    emp_wb.save(EMP_OUTPUT)
    dept_wb.save(DEPT_DESKTOP)
    emp_wb.save(EMP_DESKTOP)

    print(f'departments={len(departments)} employees={len(employees)}')
    print(DEPT_OUTPUT)
    print(EMP_OUTPUT)
    print(DEPT_DESKTOP)
    print(EMP_DESKTOP)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
