#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import re
import sys
import openpyxl
from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parent.parent

SOURCE_XLSX = Path('/mnt/c/Users/Administrator/Desktop/New folder (2)/assets_export (6).xlsx')
OUTPUT_XLSX = REPO_ROOT / 'state' / 'assets_clean_for_import.xlsx'

IMPORT_HEADERS = [
    'Mã thiết bị', 'Tên thiết bị', 'Loại thiết bị', 'Model', 'Serial', 'IP',
    'Bộ phận', 'Người dùng', 'Vị trí', 'Ngày mua', 'Hết bảo hành', 'Trạng thái', 'Ghi chú'
]

DEPT_MAP = {
    'Kế Toán': 'Kế Toán', 'kế toán': 'Kế Toán',
    'P.Tổng Hợp': 'P.Tổng Hợp',
    'P.Sales Ex': 'Sales', 'Sales': 'Sales',
    'P.ISO': 'P.ISO',
    'Kho FG': 'Kho FG',
    'XNK': 'XNK',
    'Thu Mua': 'Thu Mua',
    'HR': 'HR',
    'IT': 'IT',
    'P.Sản Xuất': 'P.Sản Xuất', 'P.SX': 'P.Sản Xuất', 'P2 P.SX': 'P.Sản Xuất',
    'Thổi Hủ': 'P.Sản Xuất', 'P.Thổi hủ': 'P.Sản Xuất', 'Ép nhựa': 'P.Sản Xuất',
    'Vô Hủ': 'P.Sản Xuất', 'Vô Hủ 1': 'P.Sản Xuất', 'Vô Hủ 2': 'P.Sản Xuất',
    'Chiên': 'P.Sản Xuất', 'Khu Chiên': 'P.Sản Xuất',
    'Sấy 1-2': 'P.Sản Xuất', 'Sấy 3-4': 'P.Sản Xuất', 'Khu Sấy Vỏ lụa': 'P.Sản Xuất',
    'P.Đóng Gói': 'P.Sản Xuất', 'P.Đóng Gói 02': 'P.Sản Xuất', 'P.Đóng gói 1': 'P.Sản Xuất',
    'Đóng goi 1': 'P.Sản Xuất', 'P2 Đóng Gói/ Chiên': 'P.Sản Xuất', 'P2-P.Đóng gói hạt dẻ': 'P.Sản Xuất',
    'P2 .Đóng Gói/ Candy': 'P.Sản Xuất', 'P2 Candy Room': 'P.Sản Xuất', 'P2 Sấy': 'P.Sản Xuất',
    'Khu Chẻ': 'P.Sản Xuất', 'Khu WIP': 'P.Sản Xuất', 'Khu Satake': 'P.Sản Xuất',
    'P1-Lon Giấy': 'P.Sản Xuất', 'P.Lon Giấy': 'P.Sản Xuất',
    'P.WH': 'Kho', 'WH': 'Kho', 'WH-SSI': 'Kho', 'P2 WHSSI': 'Kho', 'P2 Office Wh01': 'Kho', 'P2 Office Wh02': 'Kho',
    'P.Bảo Trì': 'P.Bảo Trì', 'P.Bao Trì': 'P.Bảo Trì',
    'P.QC': 'P.QC', 'P.QC-01': 'P.QC',
    'P.LAB': 'LAB', 'LAB': 'LAB',
    'Lễ Tân': 'Hành Chính', 'Office': 'Hành Chính',
    'PR Assistant': 'PR', 'PR': 'PR',
    'Y Tế': 'Y Tế', 'Y Te': 'Y Tế',
    'QA office': 'QA',
    'BV': 'An Ninh', 'BV-01': 'An Ninh', 'BV-04': 'An Ninh', 'BV-05': 'An Ninh', 'BV-06 cookies': 'An Ninh', 'BV3': 'An Ninh', 'BV4': 'An Ninh',
    'Design': 'Design',
}

LOCATION_ONLY_DEPTS = {
    'Planning Room', 'Meeting Room', 'Show Room', 'Showroom', 'New York Room', 'Canada Room', 'Free Room', 'Office', 'Salon', 'Trạm Bình Phước'
}

SHARED_USER_VALUES = {
    '', 'None', 'none', 'Free', 'All user', 'all user', 'Tổ Trưởng', 'Tổ trưởng', 'QC', 'HR', 'WH-PL', 'C2', 'BV', 'Bảo trì', 'Y Tế', 'Y Te', 'HP PC', '( HP Laptop )'
}

TYPE_MAP = {
    'PC': 'PC',
    'Printer': 'Printer',
    'NVR': 'NVR',
}


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_site(raw_location: str) -> str:
    value = clean_text(raw_location)
    if not value or value.lower() == 'none':
        return ''
    upper = value.upper()
    if upper == 'CCTV PC':
        return 'DOF'
    return upper


def titleish(value: str) -> str:
    return re.sub(r'\s+', ' ', value).strip()


def normalize_location(site: str, raw_department: str, raw_location: str) -> str:
    dept = clean_text(raw_department)
    location = clean_text(raw_location)
    location = '' if location.lower() == 'none' else location

    if dept in {'P.QC-01'}:
        area = 'QC-01'
    elif dept in {'P2 Candy Room'}:
        area = 'Candy Room'
    elif dept in {'P2 Sấy'}:
        area = 'Sấy'
    elif dept in {'P2-P.Đóng gói hạt dẻ'}:
        area = 'Đóng Gói Hạt Dẻ'
    elif dept in {'Show Room', 'Showroom'}:
        area = 'Showroom'
    elif dept in LOCATION_ONLY_DEPTS:
        area = dept
    elif dept in {'Thổi Hủ', 'P.Thổi hủ', 'Ép nhựa', 'Vô Hủ', 'Vô Hủ 1', 'Vô Hủ 2', 'Chiên', 'Khu Chiên', 'Sấy 1-2', 'Sấy 3-4', 'Khu Sấy Vỏ lụa', 'P.Đóng Gói', 'P.Đóng Gói 02', 'P.Đóng gói 1', 'Đóng goi 1', 'Khu Chẻ', 'Khu WIP', 'Khu Satake', 'P1-Lon Giấy', 'P.Lon Giấy'}:
        area = dept
    else:
        area = dept if dept else location

    area = titleish(area)
    if not site and location and location.upper() != 'CCTV PC':
        return location.upper()
    if not site:
        return area
    if not area:
        return site
    if area.upper() == site.upper():
        return site
    return f'{site} - {area}'


def normalize_department(raw_department: str) -> str:
    dept = clean_text(raw_department)
    if not dept:
        return ''
    return DEPT_MAP.get(dept, dept)


def normalize_assigned_user(raw_user: str) -> tuple[str, str | None]:
    user = clean_text(raw_user)
    if user in SHARED_USER_VALUES:
        if user:
            return '', f'Nhãn người dùng cũ: {user}'
        return '', None
    return user, None


def normalize_status(raw_status: str, assigned_user: str) -> str:
    status = clean_text(raw_status).lower()
    if status == 'active':
        return 'assigned' if assigned_user else 'in_stock'
    if status in {'assigned', 'borrowed', 'repairing', 'retired', 'disposed', 'lost', 'in_stock'}:
        return status
    return 'assigned' if assigned_user else 'in_stock'


def main() -> int:
    if not SOURCE_XLSX.exists():
        raise SystemExit(f'Không tìm thấy file nguồn: {SOURCE_XLSX}')

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    header_idx = {name: idx for idx, name in enumerate(headers)}

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'AssetsClean'
    out_ws.append(IMPORT_HEADERS)

    review_ws = out_wb.create_sheet('Review')
    review_ws.append([
        'Mã thiết bị', 'Tên thiết bị', 'Department gốc', 'Department chuẩn', 'Location gốc', 'Location chuẩn',
        'Assigned user gốc', 'Assigned user chuẩn', 'Status gốc', 'Status chuẩn', 'Ghi chú review'
    ])

    for row in data_rows:
        asset_code = clean_text(row[header_idx['Mã thiết bị']])
        asset_name = clean_text(row[header_idx['Tên thiết bị']]) or asset_code
        asset_type_raw = clean_text(row[header_idx['Loại thiết bị']])
        model = clean_text(row[header_idx['Model']])
        serial = clean_text(row[header_idx['Serial']])
        ip = clean_text(row[header_idx['IP']])
        dept_raw = clean_text(row[header_idx['Bộ phận']])
        user_raw = clean_text(row[header_idx['Người dùng']])
        location_raw = clean_text(row[header_idx['Vị trí']])
        purchase_date = clean_text(row[header_idx['Ngày mua']])
        warranty_expiry = clean_text(row[header_idx['Hết bảo hành']])
        status_raw = clean_text(row[header_idx['Trạng thái']])
        notes_raw = clean_text(row[header_idx['Ghi chú']])

        asset_type = TYPE_MAP.get(asset_type_raw, asset_type_raw or 'PC')
        site = normalize_site(location_raw)
        department = normalize_department(dept_raw)
        assigned_user, user_note = normalize_assigned_user(user_raw)
        location = normalize_location(site, dept_raw, location_raw)
        status = normalize_status(status_raw, assigned_user)

        notes = []
        if notes_raw:
            notes.append(notes_raw)
        if user_note:
            notes.append(user_note)
        if dept_raw in LOCATION_ONLY_DEPTS:
            notes.append(f'Đã chuyển khu vực từ Bộ phận sang Vị trí: {dept_raw}')
        if location_raw.strip().lower() == 'none':
            notes.append('Vị trí gốc là None')
        if not department:
            notes.append('Cần xác nhận Department')
        if not location:
            notes.append('Cần xác nhận Location')

        note_text = ' | '.join(dict.fromkeys([n for n in notes if n]))

        out_ws.append([
            asset_code, asset_name, asset_type, model or None, serial or None, ip or None,
            department or None, assigned_user or None, location or None,
            purchase_date or None, warranty_expiry or None, status, note_text or None,
        ])

        needs_review = (
            not department or not location or dept_raw in LOCATION_ONLY_DEPTS or location_raw.strip().lower() == 'none'
            or user_raw in SHARED_USER_VALUES or dept_raw not in DEPT_MAP and bool(dept_raw)
        )
        if needs_review:
            review_ws.append([
                asset_code, asset_name, dept_raw, department, location_raw, location,
                user_raw, assigned_user, status_raw, status, note_text,
            ])

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUTPUT_XLSX)
    print(f'Wrote: {OUTPUT_XLSX}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
