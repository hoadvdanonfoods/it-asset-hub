from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.auth import require_permission
from app.db.models import Asset, AssetAssignment, User
from app.db.models.master_data import AssetType, Department, Employee, Location
from app.security import hash_password
from app.db.models.master_reference import AssetCategory, AssetStatus, Vendor, IncidentCategory, Priority, MaintenanceType
from app.db.session import get_db
from app.services.audit import log_audit

import base64
import io
import json
import openpyxl

router = APIRouter(prefix='/master-data', tags=['master_data'])
templates = Jinja2Templates(directory='app/templates')


def _md_preview_token(rows_cleaned: list[dict], filename: str) -> str:
    payload = {'filename': filename, 'rows': rows_cleaned}
    return base64.urlsafe_b64encode(json.dumps(payload, ensure_ascii=False).encode('utf-8')).decode('ascii')


def _rows_from_md_token(token: str) -> tuple[list[dict], str]:
    decoded = base64.urlsafe_b64decode(token.encode('ascii'))
    payload = json.loads(decoded.decode('utf-8'))
    return payload['rows'], payload.get('filename', 'import.xlsx')


def _build_md_preview(rows_cleaned: list[dict], config: dict, db: Session, filename: str) -> dict:
    unique_field = config.get('unique_field')
    created = updated = 0
    sample_rows = []
    for i, cleaned in enumerate(rows_cleaned):
        unique_value = cleaned.get(unique_field) if unique_field else None
        if unique_field and unique_value not in (None, ''):
            existing = db.scalar(select(config['model']).where(getattr(config['model'], unique_field) == unique_value))
            action = 'update' if existing else 'create'
        else:
            action = 'create'
        if action == 'create':
            created += 1
        else:
            updated += 1
        name_val = cleaned.get('name') or cleaned.get('full_name') or ''
        sample_rows.append({
            'row_number': i + 2,
            'code': cleaned.get(unique_field, '') if unique_field else '',
            'name': name_val,
            'action': action,
        })
    return {
        'filename': filename,
        'created': created,
        'updated': updated,
        'total_rows': created + updated,
        'sample_rows': sample_rows,
        'token': _md_preview_token(rows_cleaned, filename),
    }

def _auto_create_user_for_employee(employee, db: Session) -> bool:
    code = getattr(employee, 'employee_code', None)
    if not code:
        return False
    existing = db.scalar(select(User).where(User.username == code))
    if existing:
        return False
    db.add(User(
        username=code,
        password=code,
        password_hash=hash_password(code),
        role='user',
        full_name=getattr(employee, 'full_name', None),
        is_active=True,
        must_change_password=True,
        can_view_dashboard=True,
        can_view_assets=False,
        can_view_maintenance=False,
        can_view_incidents=True,
        can_view_resources=False,
        can_create_assets=False,
        can_edit_assets=False,
        can_import_assets=False,
        can_export_assets=False,
        can_create_maintenance=False,
        can_edit_maintenance=False,
        can_export_maintenance=False,
        can_create_incidents=True,
        can_edit_incidents=False,
        can_export_incidents=False,
        can_manage_users=False,
        can_manage_system=False,
        can_manage_resources=False,
        can_view_documents=False,
        can_manage_documents=False,
    ))
    return True


MODEL_CONFIG = {
    'departments': {
        'label': 'PhÃ²ng ban',
        'description': 'Quáº£n lÃ½ danh má»¥c phÃ²ng ban dÃ¹ng chung cho tÃ i sáº£n, nhÃ¢n sá»± vÃ  bÃ¡o cÃ¡o.',
        'icon': 'building',
        'accent': 'blue',
        'model': Department,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£', 'type': 'text', 'required': True, 'placeholder': 'VD: IT'},
            {'key': 'name', 'label': 'TÃªn phÃ²ng ban', 'type': 'text', 'required': True, 'placeholder': 'VD: PhÃ²ng CÃ´ng nghá»‡ ThÃ´ng tin'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'note', 'label': 'Ghi chÃº', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
        ],
        'table_columns': ['code', 'name', 'is_active', 'note'],
        'unique_field': 'code',
    },
    'employees': {
        'label': 'NhÃ¢n viÃªn',
        'description': 'Quáº£n lÃ½ danh má»¥c nhÃ¢n sá»± cÆ¡ báº£n phá»¥c vá»¥ bÃ n giao tÃ i sáº£n vÃ  tra cá»©u ná»™i bá»™.',
        'icon': 'users',
        'accent': 'emerald',
        'model': Employee,
        'title_field': 'full_name',
        'columns': [
            {'key': 'employee_code', 'label': 'MÃ£ nhÃ¢n viÃªn', 'type': 'text', 'required': True, 'placeholder': 'VD: NV001'},
            {'key': 'full_name', 'label': 'Há» vÃ  tÃªn', 'type': 'text', 'required': True, 'placeholder': 'Nguyá»…n VÄƒn A'},
            {'key': 'department_id', 'label': 'PhÃ²ng ban', 'type': 'select', 'placeholder': 'Chá»n phÃ²ng ban', 'options_source': 'departments', 'option_value': 'id', 'option_label': 'name'},
            {'key': 'title', 'label': 'Chá»©c danh', 'type': 'text', 'placeholder': 'VD: IT Support'},
            {'key': 'email', 'label': 'Email', 'type': 'email', 'placeholder': 'name@company.com'},
            {'key': 'phone', 'label': 'Sá»‘ Ä‘iá»‡n thoáº¡i', 'type': 'text', 'placeholder': '090xxxxxxx'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'note', 'label': 'Ghi chÃº', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
        ],
        # import_columns: dÃ¹ng riÃªng cho file Excel import/export (dÃ¹ng mÃ£ PB thay vÃ¬ ID)
        'import_columns': [
            {'key': 'employee_code', 'label': 'MÃ£ nhÃ¢n viÃªn', 'type': 'text', 'required': True, 'placeholder': 'VD: NV001'},
            {'key': 'full_name', 'label': 'Há» vÃ  tÃªn', 'type': 'text', 'required': True, 'placeholder': 'Nguyá»…n VÄƒn A'},
            {'key': 'department_code', 'label': 'MÃ£ phÃ²ng ban', 'type': 'text', 'placeholder': 'VD: IT'},
            {'key': 'department_id', 'label': 'ID phÃ²ng ban (TÃ¹y chá»n)', 'type': 'number', 'placeholder': 'VD: 1'},
            {'key': 'title', 'label': 'Chá»©c danh', 'type': 'text', 'placeholder': 'VD: IT Support'},
            {'key': 'email', 'label': 'Email', 'type': 'email', 'placeholder': 'name@company.com'},
            {'key': 'phone', 'label': 'Sá»‘ Ä‘iá»‡n thoáº¡i', 'type': 'text', 'placeholder': '090xxxxxxx'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'note', 'label': 'Ghi chÃº', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
        ],
        'table_columns': ['employee_code', 'full_name', 'department_id', 'title', 'email', 'phone', 'is_active'],
        'table_display': {'department_id': 'department_name'},
        'unique_field': 'employee_code',
    },
    'asset_types': {
        'label': 'Loáº¡i tÃ i sáº£n',
        'description': 'Quáº£n lÃ½ danh má»¥c loáº¡i tÃ i sáº£n, nhÃ³m phÃ¢n loáº¡i vÃ  quy Æ°á»›c Ä‘áº·t mÃ£.',
        'icon': 'tag',
        'accent': 'amber',
        'model': AssetType,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ loáº¡i', 'type': 'text', 'required': True, 'placeholder': 'VD: LAPTOP'},
            {'key': 'name', 'label': 'TÃªn loáº¡i tÃ i sáº£n', 'type': 'text', 'required': True, 'placeholder': 'VD: Laptop'},
            {'key': 'category_group', 'label': 'NhÃ³m phÃ¢n loáº¡i', 'type': 'text', 'placeholder': 'VD: User Device'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'note', 'label': 'Ghi chÃº', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
        ],
        'table_columns': ['code', 'name', 'category_group', 'is_active', 'note'],
        'unique_field': 'code',
    },
    'locations': {
        'label': 'Vá»‹ trÃ­',
        'description': 'Quáº£n lÃ½ danh má»¥c vá»‹ trÃ­, khu vá»±c vÃ  Ä‘iá»ƒm Ä‘áº·t tÃ i sáº£n trong há»‡ thá»‘ng.',
        'icon': 'map-pin',
        'accent': 'rose',
        'model': Location,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ vá»‹ trÃ­', 'type': 'text', 'required': True, 'placeholder': 'VD: HN-F1'},
            {'key': 'name', 'label': 'TÃªn vá»‹ trÃ­', 'type': 'text', 'required': True, 'placeholder': 'VD: HÃ  Ná»™i - Táº§ng 1'},
            {'key': 'site_group', 'label': 'NhÃ³m khu vá»±c', 'type': 'text', 'placeholder': 'VD: Head Office'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'note', 'label': 'Ghi chÃº', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
        ],
        'table_columns': ['code', 'name', 'site_group', 'is_active', 'note'],
        'unique_field': 'code',
    },
    'asset_categories': {
        'label': 'NhÃ³m tÃ i sáº£n',
        'description': 'Danh má»¥c chuáº©n hÃ³a nhÃ³m tÃ i sáº£n cho Phase 1 normalization.',
        'icon': 'boxes',
        'accent': 'violet',
        'model': AssetCategory,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ nhÃ³m', 'type': 'text', 'required': True, 'placeholder': 'VD: LAPTOP'},
            {'key': 'name', 'label': 'TÃªn nhÃ³m', 'type': 'text', 'required': True, 'placeholder': 'VD: Laptop'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
    'asset_statuses': {
        'label': 'Tráº¡ng thÃ¡i tÃ i sáº£n',
        'description': 'Danh má»¥c tráº¡ng thÃ¡i tÃ i sáº£n chuáº©n hÃ³a dÃ¹ng chung.',
        'icon': 'activity',
        'accent': 'emerald',
        'model': AssetStatus,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ tráº¡ng thÃ¡i', 'type': 'text', 'required': True, 'placeholder': 'VD: IN_STOCK'},
            {'key': 'name', 'label': 'TÃªn tráº¡ng thÃ¡i', 'type': 'text', 'required': True, 'placeholder': 'VD: In Stock'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
    'vendors': {
        'label': 'NhÃ  cung cáº¥p',
        'description': 'Danh má»¥c nhÃ  cung cáº¥p vÃ  Ä‘á»‘i tÃ¡c báº£o trÃ¬.',
        'icon': 'truck',
        'accent': 'amber',
        'model': Vendor,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ NCC', 'type': 'text', 'required': True, 'placeholder': 'VD: DELL'},
            {'key': 'name', 'label': 'TÃªn NCC', 'type': 'text', 'required': True, 'placeholder': 'VD: Dell Viá»‡t Nam'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
    'incident_categories': {
        'label': 'NhÃ³m sá»± cá»‘',
        'description': 'Danh má»¥c phÃ¢n loáº¡i ticket sá»± cá»‘.',
        'icon': 'shield-alert',
        'accent': 'rose',
        'model': IncidentCategory,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ nhÃ³m', 'type': 'text', 'required': True, 'placeholder': 'VD: HARDWARE'},
            {'key': 'name', 'label': 'TÃªn nhÃ³m', 'type': 'text', 'required': True, 'placeholder': 'VD: Hardware'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
    'priorities': {
        'label': 'Äá»™ Æ°u tiÃªn',
        'description': 'Danh má»¥c Æ°u tiÃªn chuáº©n hÃ³a cho ticket sá»± cá»‘.',
        'icon': 'flame',
        'accent': 'orange',
        'model': Priority,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ Æ°u tiÃªn', 'type': 'text', 'required': True, 'placeholder': 'VD: HIGH'},
            {'key': 'name', 'label': 'TÃªn Æ°u tiÃªn', 'type': 'text', 'required': True, 'placeholder': 'VD: High'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
    'maintenance_types': {
        'label': 'Loáº¡i báº£o trÃ¬',
        'description': 'Danh má»¥c loáº¡i báº£o trÃ¬ chuáº©n hÃ³a.',
        'icon': 'wrench',
        'accent': 'yellow',
        'model': MaintenanceType,
        'title_field': 'name',
        'columns': [
            {'key': 'code', 'label': 'MÃ£ loáº¡i', 'type': 'text', 'required': True, 'placeholder': 'VD: PREVENTIVE'},
            {'key': 'name', 'label': 'TÃªn loáº¡i', 'type': 'text', 'required': True, 'placeholder': 'VD: Preventive'},
            {'key': 'description', 'label': 'MÃ´ táº£', 'type': 'textarea', 'placeholder': 'ThÃ´ng tin bá»• sung...'},
            {'key': 'is_active', 'label': 'KÃ­ch hoáº¡t', 'type': 'boolean'},
            {'key': 'sort_order', 'label': 'Thá»© tá»±', 'type': 'number', 'placeholder': '0'},
        ],
        'table_columns': ['code', 'name', 'description', 'is_active', 'sort_order'],
        'unique_field': 'code',
    },
}


def _get_config(model_name: str):
    return MODEL_CONFIG.get(model_name)


def _get_form_options(db: Session, config):
    options = {}
    sources = {
        'departments': db.scalars(select(Department).where(Department.is_active == True).order_by(Department.name.asc())).all(),  # noqa: E712
        'employees': db.scalars(select(Employee).where(Employee.is_active == True).order_by(Employee.full_name.asc())).all(),  # noqa: E712
        'locations': db.scalars(select(Location).where(Location.is_active == True).order_by(Location.name.asc())).all(),  # noqa: E712
        'asset_types': db.scalars(select(AssetType).where(AssetType.is_active == True).order_by(AssetType.name.asc())).all(),  # noqa: E712
        'asset_categories': db.scalars(select(AssetCategory).where(AssetCategory.is_active == True).order_by(AssetCategory.name.asc())).all(),  # noqa: E712
        'asset_statuses': db.scalars(select(AssetStatus).where(AssetStatus.is_active == True).order_by(AssetStatus.sort_order.asc(), AssetStatus.name.asc())).all(),  # noqa: E712
        'vendors': db.scalars(select(Vendor).where(Vendor.is_active == True).order_by(Vendor.name.asc())).all(),  # noqa: E712
        'priorities': db.scalars(select(Priority).where(Priority.is_active == True).order_by(Priority.sort_order.asc(), Priority.name.asc())).all(),  # noqa: E712
        'incident_categories': db.scalars(select(IncidentCategory).where(IncidentCategory.is_active == True).order_by(IncidentCategory.name.asc())).all(),  # noqa: E712
        'maintenance_types': db.scalars(select(MaintenanceType).where(MaintenanceType.is_active == True).order_by(MaintenanceType.name.asc())).all(),  # noqa: E712
    }
    for field in config['columns']:
        source_name = field.get('options_source')
        if source_name:
            options[field['key']] = sources.get(source_name, [])
    return options


def _parse_form_data(request_form, config):
    data = {}
    for field in config['columns']:
        key = field['key']
        if field['type'] == 'boolean':
            data[key] = request_form.get(key) == 'true'
            continue
        raw = request_form.get(key)
        value = raw.strip() if isinstance(raw, str) else raw
        if value == '':
            value = None
        if field['type'] == 'number' and value is not None:
            try:
                value = int(value)
            except (TypeError, ValueError):
                value = None
        data[key] = value
    return data


def _clean_import_row(row_data, config):
    allowed_fields = {field['key'] for field in config['columns'] if field['key'] != 'id'}
    cleaned = {}
    for field in config['columns']:
        key = field['key']
        if key not in allowed_fields:
            continue
        value = row_data.get(key)
        if value is not None and field.get('type') in ('text', 'email', 'textarea'):
            value = str(value)

        if isinstance(value, str):
            value = value.strip()
        if value == '':
            value = None
        if field['type'] == 'boolean':
            value = str(value).strip().lower() in ('1', 'true', 'yes', 'y', 'on') if value is not None else False
        if field['type'] == 'number' and value is not None:
            try:
                value = int(value)
            except (TypeError, ValueError):
                value = None
        cleaned[key] = value
    return cleaned


# Map FK field -> (model, pk column) for FK validation during import
_FK_FIELD_MODELS = {
    'department_id': (Department, 'id'),
}

# Fallback lookup fields when the integer ID is not found (allows importing by code or name)
_FK_FIELD_LOOKUP_FALLBACK = {
    'department_id': ['code', 'name'],
}


def _sanitize_fk_fields(cleaned: dict, db: Session) -> dict:
    """Resolve FK fields by ID first, then by code/name if ID not found."""
    for field_key, (model, pk_col) in _FK_FIELD_MODELS.items():
        if field_key not in cleaned:
            continue
        raw_id = cleaned[field_key]
        if raw_id is None:
            continue

        # Try direct ID lookup
        exists = db.scalar(select(model).where(getattr(model, pk_col) == raw_id))
        if exists:
            continue

        # Try fallback lookup by code/name (handles string values or cross-DB imports)
        fallback_fields = _FK_FIELD_LOOKUP_FALLBACK.get(field_key, [])
        lookup_value = str(raw_id).strip() if raw_id is not None else None
        if fallback_fields and lookup_value:
            found = None
            for lookup_field in fallback_fields:
                found = db.scalar(select(model).where(getattr(model, lookup_field).ilike(lookup_value)))
                if found:
                    break
            if found:
                cleaned[field_key] = getattr(found, pk_col)
                continue

        cleaned[field_key] = None
    return cleaned


def _resolve_department_code(cleaned: dict, db: Session) -> dict:
    """Chuyá»ƒn department_code (mÃ£ phÃ²ng ban) â†’ department_id khi import nhÃ¢n viÃªn."""
    # Æ¯u tiÃªn náº¿u trong file cÃ³ cá»™t department_id (há»£p lá»‡)
    dept_id = cleaned.get('department_id')
    if dept_id is not None:
        try:
            dept = db.get(Department, int(dept_id))
            cleaned['department_id'] = dept.id if dept else None
            cleaned.pop('department_code', None)  # Bá» qua department_code
            return cleaned
        except (ValueError, TypeError):
            pass  # Náº¿u khÃ´ng parse Ä‘Æ°á»£c int thÃ¬ tiáº¿p tá»¥c dÃ¹ng department_code

    if 'department_code' not in cleaned:
        return cleaned
        
    dept_code = cleaned.pop('department_code', None)
    if dept_code:
        dept = db.scalar(select(Department).where(Department.code == dept_code))
        if dept:
            cleaned['department_id'] = dept.id
        elif str(dept_code).strip().isdigit():
            # Fallback: Náº¿u há» Ä‘iá»n sá»‘ ID vÃ o cá»™t department_code
            dept = db.get(Department, int(str(dept_code).strip()))
            cleaned['department_id'] = dept.id if dept else None
        else:
            cleaned['department_id'] = None
    else:
        cleaned.setdefault('department_id', None)
    return cleaned


@router.get('/{model_name}', response_class=HTMLResponse)
@require_permission('can_manage_system')
async def list_model(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    items = db.scalars(select(config['model']).order_by(config['model'].id.asc())).all()
    if model_name == 'employees':
        department_lookup = {dept.id: dept.name for dept in db.scalars(select(Department)).all()}
        for item in items:
            item.department_name = department_lookup.get(getattr(item, 'department_id', None), '')
    return templates.TemplateResponse(
        'master_data/list.html',
        {
            'request': request,
            'current_user': current_user,
            'model_name': model_name,
            'config': config,
            'items': items,
            'field_options': _get_form_options(db, config),
            'p': '/master-data/' + model_name,
        },
    )


@router.get('/{model_name}/bulk-edit', response_class=HTMLResponse)
@require_permission('can_manage_system')
async def bulk_edit_model(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    return RedirectResponse(f'/master-data/{model_name}', status_code=303)


@router.post('/{model_name}/create')
@require_permission('can_manage_system')
async def create_model(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    form = await request.form()
    data = _parse_form_data(form, config)
    new_obj = config['model'](**data)
    db.add(new_obj)
    if model_name == 'employees':
        db.flush()
        _auto_create_user_for_employee(new_obj, db)
    db.commit()
    return RedirectResponse(f'/master-data/{model_name}', status_code=303)


@router.post('/{model_name}/edit/{item_id}')
@require_permission('can_manage_system')
async def edit_model(request: Request, model_name: str, item_id: int, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    item = db.get(config['model'], item_id)
    if not item:
        return RedirectResponse(f'/master-data/{model_name}', status_code=303)
    form = await request.form()
    data = _parse_form_data(form, config)
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    return RedirectResponse(f'/master-data/{model_name}', status_code=303)


@router.post('/{model_name}/bulk-edit')
@require_permission('can_manage_system')
async def bulk_edit_model_submit(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)

    form = await request.form()
    item_ids_raw = (form.get('item_ids') or '').strip()
    if not item_ids_raw:
        return RedirectResponse(f'/master-data/{model_name}', status_code=303)

    item_ids = []
    for token in item_ids_raw.split(','):
        token = token.strip()
        if token.isdigit():
            item_ids.append(int(token))

    items = db.scalars(select(config['model']).where(config['model'].id.in_(item_ids))).all() if item_ids else []

    updates = {}
    for field in config['columns']:
        key = field['key']
        raw = form.get(key)
        if field['type'] == 'boolean':
            if raw in (None, ''):
                continue
            updates[key] = str(raw).strip().lower() == 'true'
            continue

        value = raw.strip() if isinstance(raw, str) else raw
        if value in (None, ''):
            continue
        if field['type'] == 'number':
            try:
                value = int(value)
            except (TypeError, ValueError):
                continue
        updates[key] = value

    for item in items:
        for key, value in updates.items():
            setattr(item, key, value)

    db.commit()
    return RedirectResponse(f'/master-data/{model_name}', status_code=303)


def _redirect_with_bulk_feedback(model_name: str, *, message: str | None = None, success: int = 0, blocked: int = 0, details: list[str] | None = None, tone: str = 'info'):
    params = [f'bulk_tone={tone}']
    if message:
        params.append(f'bulk_message={message}')
    if success:
        params.append(f'bulk_success={success}')
    if blocked:
        params.append(f'bulk_blocked={blocked}')
    if details:
        params.append(f"bulk_details={' || '.join(details[:12])}")
    query = '&'.join(params)
    return RedirectResponse(f"/master-data/{model_name}{'?' + query if query else ''}", status_code=303)


@router.post('/{model_name}/bulk-archive')
@require_permission('can_manage_system')
async def bulk_archive_model_submit(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)

    form = await request.form()
    item_ids_raw = (form.get('item_ids') or '').strip()
    confirm_text = (form.get('confirm_text') or '').strip().upper()
    if not item_ids_raw or confirm_text != 'ARCHIVE':
        return _redirect_with_bulk_feedback(model_name, message='XÃ¡c nháº­n khÃ´ng há»£p lá»‡')

    item_ids = []
    for token in item_ids_raw.split(','):
        token = token.strip()
        if token.isdigit():
            item_ids.append(int(token))

    success = 0
    blocked = 0
    details: list[str] = []

    if item_ids:
        items = db.scalars(select(config['model']).where(config['model'].id.in_(item_ids))).all()
        for item in items:
            if not hasattr(item, 'is_active'):
                blocked += 1
                continue

            if model_name == 'employees':
                active_assets = db.scalars(
                    select(Asset).where(
                        or_(Asset.current_assignment_id.is_not(None), Asset.assigned_user == item.full_name)
                    )
                ).all()
                blocking_assets = []
                for asset in active_assets:
                    active_assignment = db.scalar(
                        select(AssetAssignment).where(
                            AssetAssignment.asset_id == asset.id,
                            AssetAssignment.unassigned_at.is_(None),
                            AssetAssignment.employee_id == item.id,
                        ).order_by(AssetAssignment.assigned_at.desc())
                    )
                    if active_assignment or (asset.assigned_user and asset.assigned_user == item.full_name):
                        blocking_assets.append(asset.asset_code)
                if blocking_assets:
                    blocked += 1
                    preview = ', '.join(blocking_assets[:3])
                    suffix = '' if len(blocking_assets) <= 3 else f' vÃ  {len(blocking_assets) - 3} tÃ i sáº£n khÃ¡c'
                    details.append(f'{item.full_name} ({item.employee_code}): cÃ²n giá»¯ {len(blocking_assets)} tÃ i sáº£n, gá»“m {preview}{suffix}')
                    log_audit(db, actor=current_user.username if current_user else None, module='master_data', action='bulk_archive', entity_type='employee', entity_id=item.id, result='skipped', reason='active_assets', metadata={'employee_code': item.employee_code, 'assets': blocking_assets[:10]})
                    continue

            if getattr(item, 'is_active', None) is False:
                blocked += 1
                details.append(f'{getattr(item, config["title_field"], item.id)}: Ä‘Ã£ inactive')
                continue

            setattr(item, 'is_active', False)
            success += 1
            log_audit(db, actor=current_user.username if current_user else None, module='master_data', action='bulk_archive', entity_type=model_name.rstrip('s'), entity_id=item.id, metadata={'model': model_name, 'title': getattr(item, config['title_field'], None)})

    db.commit()
    tone = 'success' if success and not blocked else 'warning' if blocked else 'info'
    return _redirect_with_bulk_feedback(model_name, message='ÄÃ£ xá»­ lÃ½ inactive hÃ ng loáº¡t', success=success, blocked=blocked, details=details, tone=tone)


@router.post('/{model_name}/bulk-delete')
@require_permission('can_manage_system')
async def bulk_delete_model_submit(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)

    form = await request.form()
    item_ids_raw = (form.get('item_ids') or '').strip()
    confirm_text = (form.get('confirm_text') or '').strip().upper()
    if not item_ids_raw or confirm_text != 'DELETE':
        return _redirect_with_bulk_feedback(model_name, message='XÃ¡c nháº­n khÃ´ng há»£p lá»‡. Nháº­p chá»¯ DELETE Ä‘á»ƒ tiáº¿p tá»¥c.', tone='warning')

    item_ids = [int(t) for t in item_ids_raw.split(',') if t.strip().isdigit()]
    success = 0
    blocked = 0
    details: list[str] = []

    if item_ids:
        items = db.scalars(select(config['model']).where(config['model'].id.in_(item_ids))).all()
        for item in items:
            title = str(getattr(item, config['title_field'], item.id))

            # Block employees with active asset assignments
            if model_name == 'employees':
                active_assignment = db.scalar(
                    select(AssetAssignment).where(
                        AssetAssignment.employee_id == item.id,
                        AssetAssignment.unassigned_at.is_(None),
                    ).limit(1)
                )
                if active_assignment:
                    blocked += 1
                    details.append(f'{title}: cÃ²n bÃ n giao tÃ i sáº£n chÆ°a thu há»“i')
                    continue

            try:
                db.delete(item)
                db.flush()
                success += 1
                log_audit(db, actor=current_user.username if current_user else None, module='master_data', action='bulk_delete', entity_type=model_name, entity_id=item.id, metadata={'title': title})
            except IntegrityError:
                db.rollback()
                blocked += 1
                details.append(f'{title}: Ä‘ang Ä‘Æ°á»£c tham chiáº¿u bá»Ÿi dá»¯ liá»‡u khÃ¡c')

    db.commit()
    tone = 'success' if success and not blocked else 'warning' if blocked else 'info'
    return _redirect_with_bulk_feedback(model_name, message=f'ÄÃ£ xÃ³a {success} báº£n ghi.', success=success, blocked=blocked, details=details, tone=tone)


@router.post('/employees/bulk-create-users')
@require_permission('can_manage_system')
async def bulk_create_users_for_employees(request: Request, current_user=None, db: Session = Depends(get_db)):
    employees = db.scalars(select(Employee).where(Employee.is_active == True)).all()  # noqa: E712
    existing_usernames = set(db.scalars(select(User.username)).all())
    created = 0
    skipped = 0
    for emp in employees:
        if emp.employee_code and emp.employee_code not in existing_usernames:
            _auto_create_user_for_employee(emp, db)
            existing_usernames.add(emp.employee_code)
            created += 1
        else:
            skipped += 1
    db.commit()
    from urllib.parse import quote
    msg = quote(f'ÄÃ£ táº¡o {created} tÃ i khoáº£n má»›i. {skipped} nhÃ¢n viÃªn Ä‘Ã£ cÃ³ tÃ i khoáº£n hoáº·c thiáº¿u mÃ£.')
    return RedirectResponse(f'/master-data/employees?bulk_message={msg}&bulk_tone=success', status_code=303)


@router.get('/{model_name}/import', response_class=HTMLResponse)
@require_permission('can_manage_system')
async def import_model_page(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    return templates.TemplateResponse('master_data/import.html', {
        'request': request, 'current_user': current_user,
        'model_name': model_name, 'config': config,
    })


@router.post('/{model_name}/import/preview', response_class=HTMLResponse)
@require_permission('can_manage_system')
async def import_model_preview(request: Request, model_name: str, current_user=None, file: UploadFile = File(...), db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    # DÃ¹ng import_columns náº¿u cÃ³ (há»— trá»£ department_code cho employees)
    import_cols = config.get('import_columns', config['columns'])
    parse_config = {**config, 'columns': import_cols}
    rows_cleaned = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cleaned = _clean_import_row(dict(zip(headers, row)), parse_config)
        if model_name == 'employees':
            cleaned = _resolve_department_code(cleaned, db)
        if any(v is not None and v != '' for v in cleaned.values()):
            rows_cleaned.append(cleaned)
    preview = _build_md_preview(rows_cleaned, config, db, file.filename or 'import.xlsx')
    return templates.TemplateResponse('master_data/import.html', {
        'request': request, 'current_user': current_user,
        'model_name': model_name, 'config': config, 'preview': preview,
    })


@router.post('/{model_name}/import/confirm')
@require_permission('can_manage_system')
async def import_model_confirm(request: Request, model_name: str, current_user=None, preview_token: str = Form(...), db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)
    rows_cleaned, _ = _rows_from_md_token(preview_token)
    unique_field = config.get('unique_field')
    created = updated = 0
    for cleaned in rows_cleaned:
        # Validate FK references â€” nullify any IDs that don't exist in target DB
        cleaned = _sanitize_fk_fields(cleaned, db)
        unique_value = cleaned.get(unique_field) if unique_field else None
        existing = None
        if unique_field and unique_value not in (None, ''):
            existing = db.scalar(select(config['model']).where(getattr(config['model'], unique_field) == unique_value))
        if existing:
            for key, value in cleaned.items():
                setattr(existing, key, value)
            updated += 1
        else:
            new_obj = config['model'](**cleaned)
            db.add(new_obj)
            if model_name == 'employees':
                db.flush()
                _auto_create_user_for_employee(new_obj, db)
            created += 1
    db.commit()
    from urllib.parse import quote
    msg = quote(f'Import xong: {created} thÃªm má»›i, {updated} cáº­p nháº­t')
    return RedirectResponse(f'/master-data/{model_name}?bulk_message={msg}&bulk_tone=success', status_code=303)


@router.post('/{model_name}/import')
@require_permission('can_manage_system')
async def import_model(request: Request, model_name: str, current_user=None, file: UploadFile = File(...), db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)

    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    unique_field = config.get('unique_field')
    # DÃ¹ng import_columns náº¿u cÃ³ (há»— trá»£ department_code cho employees)
    import_cols = config.get('import_columns', config['columns'])
    parse_config = {**config, 'columns': import_cols}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        cleaned = _clean_import_row(row_data, parse_config)
        if model_name == 'employees':
            cleaned = _resolve_department_code(cleaned, db)
        if any(v is not None and v != '' for v in cleaned.values()):
            # Validate FK references â€” nullify any IDs that don't exist in target DB
            cleaned = _sanitize_fk_fields(cleaned, db)
            existing = None
            unique_value = cleaned.get(unique_field) if unique_field else None
            if unique_field and unique_value not in (None, ''):
                existing = db.scalar(select(config['model']).where(getattr(config['model'], unique_field) == unique_value))
            if existing:
                for key, value in cleaned.items():
                    setattr(existing, key, value)
            else:
                db.add(config['model'](**cleaned))

    db.commit()
    return RedirectResponse(f'/master-data/{model_name}', status_code=303)


@router.get('/{model_name}/export')
@require_permission('can_manage_system')
async def export_model(request: Request, model_name: str, current_user=None, db: Session = Depends(get_db)):
    config = _get_config(model_name)
    if not config:
        return RedirectResponse('/', status_code=303)

    items = db.scalars(select(config['model']).order_by(config['model'].id.asc())).all()

    # Build FK â†’ display value lookup so exported files are portable across servers
    fk_display: dict[str, dict] = {}
    if model_name == 'employees':
        depts = db.scalars(select(Department)).all()
        fk_display['department_code'] = {d.id: (d.code or d.name) for d in depts}

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = config['label'][:31]

    # DÃ¹ng import_columns náº¿u cÃ³ Ä‘á»ƒ export ra file dá»… re-import
    import_cols = config.get('import_columns', config['columns'])
    columns = [field['key'] for field in import_cols]
    sheet.append(columns)

    for item in items:
        row = []
        for key in columns:
            if key == 'department_code' and model_name == 'employees':
                dept_id = getattr(item, 'department_id', None)
                val = fk_display['department_code'].get(dept_id, '') if dept_id else ''
                row.append(val)
            else:
                value = getattr(item, key, None)
                if key in fk_display and value is not None:
                    value = fk_display[key].get(value, value)
                row.append(value)
        sheet.append(row)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f'{model_name}_export.xlsx'
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

