from collections import defaultdict
import asyncio
import os
import datetime
import openpyxl
import zipfile
import io
import json
import re
import xml.etree.ElementTree as ET
import httpx
from fastapi import APIRouter, Depends, Request, Query, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, delete
from sqlalchemy.orm import Session
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.auth import require_module_access, require_permission, get_current_user, has_permission
from app.db.session import get_db
from app.db.models import Asset, AssetEvent
from app.db.models.resource import Resource
from app.security import decrypt_resource_password

router = APIRouter(prefix="/checklist", tags=["checklist"])
templates = Jinja2Templates(directory="app/templates")


# ---------------------------------------------------------------------------
# Utility: Sort camera list by IP address numerically
# ---------------------------------------------------------------------------
def _ip_sort_key(cam):
    """Return a tuple of integers for sorting by IP address naturally."""
    ip = cam.ip_address or ""
    parts = []
    for seg in ip.split("."):
        try:
            parts.append(int(seg))
        except ValueError:
            parts.append(0)
    # Pad to 4 parts
    while len(parts) < 4:
        parts.append(0)
    return tuple(parts)


# ---------------------------------------------------------------------------
# GET /checklist/ — Main checklist page
# ---------------------------------------------------------------------------
@router.get("/", response_class=HTMLResponse)
@require_module_access("assets")
def checklist_index(request: Request, db: Session = Depends(get_db), current_user=None):
    # Danh sách tất cả NVR (đầu ghi)
    nvrs = db.scalars(select(Asset).where(Asset.asset_type == "NVR").order_by(Asset.asset_name)).all()

    # Lấy toàn bộ Camera
    cameras = db.scalars(select(Asset).where(Asset.asset_type == "Camera").order_by(Asset.asset_name)).all()

    # Phân nhóm Camera theo Location (Lưu tên hoặc IP NVR)
    grouped_cameras = defaultdict(list)
    for c in cameras:
        loc = c.location.strip() if c.location else "chưa_gán"
        grouped_cameras[loc].append(c)

    # Sort each group by IP address
    for loc in grouped_cameras:
        grouped_cameras[loc] = sorted(grouped_cameras[loc], key=_ip_sort_key)

    # Tạo map NVR objects bằng IP_ADDRESS, Tên và Mã của NVR để tăng khả năng khớp dữ liệu
    nvr_map = {}
    for n in nvrs:
        if n.ip_address:
            nvr_map[n.ip_address.strip()] = n
        if n.asset_name:
            nvr_map[n.asset_name.strip()] = n
        if n.asset_code:
            nvr_map[n.asset_code.strip()] = n

    # Current day for date picker default
    today = datetime.datetime.now().day

    return templates.TemplateResponse("assets/checklist.html", {
        "request": request,
        "nvrs": nvrs,
        "nvr_map": nvr_map,
        "grouped_cameras": grouped_cameras,
        "current_user": current_user,
        "today": today,
    })


# ---------------------------------------------------------------------------
# GET /checklist/partial — HTMX fragment for camera list by NVR
# ---------------------------------------------------------------------------
@router.get("/partial", response_class=HTMLResponse)
@require_module_access("assets")
def checklist_partial(request: Request, nvr_ip: str = Query(""), db: Session = Depends(get_db), current_user=None):
    stmt = select(Asset).where(Asset.asset_type == "Camera")
    if nvr_ip == "chưa_gán":
        stmt = stmt.where((Asset.location == None) | (Asset.location == ""))
    elif nvr_ip:
        stmt = stmt.where(Asset.location == nvr_ip)

    cameras = db.scalars(stmt).all()
    cameras = sorted(cameras, key=_ip_sort_key)

    return templates.TemplateResponse("assets/_checklist_table.html", {
        "request": request,
        "camera_list": cameras,
        "nvr_ip": nvr_ip,
        "current_user": current_user
    })


# ---------------------------------------------------------------------------
# POST /checklist/save — Save checklist evaluations + update Excel
# ---------------------------------------------------------------------------
@router.post("/save")
async def checklist_save(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return RedirectResponse(url="/login", status_code=303)

    form_data = await request.form()
    actor = current_user.username if current_user else "System"
    checked_count = 0
    excel_updates = defaultdict(list)

    # Get the selected day from form, default to today
    try:
        selected_day = int(form_data.get("selected_day", datetime.datetime.now().day))
    except (ValueError, TypeError):
        selected_day = datetime.datetime.now().day

    active_nvr = form_data.get("active_nvr", "").strip()

    for key, value in form_data.items():
        if key.startswith("cam_"):
            try:
                asset_id = int(key.split("_")[1])
                status_eval = value  # ok, warn, err
                note = form_data.get(f"notes_{asset_id}", "").strip()

                asset = db.get(Asset, asset_id)
                if not asset:
                    continue

                # NEW: Only save if this camera belongs to the active NVR tab
                if active_nvr and asset.location != active_nvr:
                    continue

                if status_eval == 'ok':
                    title = "Bảo trì: Tốt"
                    desc = note or "Bình thường (Checklist Nhanh)"
                    if asset.status == "broken":
                        asset.status = "active"
                elif status_eval == 'warn':
                    title = "Bảo trì: Cảnh báo"
                    desc = note or "Bị mờ / nhiễu / rung (Checklist Nhanh)"
                elif status_eval == 'err':
                    title = "Bảo trì: Lỗi Mất Hình"
                    desc = note or "Không lên hình / No Signal (Checklist Nhanh)"
                    asset.status = "broken"
                else:
                    continue

                # 1. LOG TO DATABASE (WITH OVERWRITE LOGIC)
                event_date = datetime.datetime.now()
                try:
                    if 1 <= selected_day <= 31:
                        event_date = event_date.replace(day=selected_day)
                except ValueError:
                    pass

                # Find and remove existing checklists for this asset on this specific day
                start_of_day = event_date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = event_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                
                db.execute(
                    delete(AssetEvent)
                    .where(AssetEvent.asset_id == asset.id)
                    .where(AssetEvent.event_type == "daily_checklist")
                    .where(AssetEvent.created_at >= start_of_day)
                    .where(AssetEvent.created_at <= end_of_day)
                )

                new_event = AssetEvent(
                    asset_id=asset.id,
                    event_type="daily_checklist",
                    title=title,
                    description=desc,
                    actor=actor,
                    created_at=event_date
                )
                db.add(new_event)

                # 2. Collect for legacy Excel sync
                nvr_ip = asset.location.strip() if asset.location else ""
                if nvr_ip:
                    excel_updates[nvr_ip].append({
                        "cam_code": asset.asset_code,
                        "cam_name": asset.asset_name,
                        "status": status_eval,
                        "note": note
                    })

                checked_count += 1
            except Exception as e:
                print(f"Error handling cam {key}: {e}")
                continue

    db.commit()

    # ------ LEGACY: Update legacy Excel Files (Best Effort) ------
    checklist_dir = "data/camera_checklists"
    if os.path.exists(checklist_dir):
        for nvr_ip, cams in excel_updates.items():
            # Find matching file
            for root, dirs, files in os.walk(checklist_dir):
                for file in files:
                    if file.endswith(".xlsx") and not file.startswith('~') and nvr_ip in file:
                        filepath = os.path.join(root, file)
                        try:
                            wb = openpyxl.load_workbook(filepath)
                            ws = wb.active
                            
                            # Find day row
                            row_idx = None
                            for r in range(10, ws.max_row + 1):
                                if str(ws.cell(row=r, column=1).value) == str(selected_day):
                                    row_idx = r
                                    break
                            
                            if row_idx:
                                # Find D-code cols
                                d_col_map = {}
                                for c in range(1, ws.max_column + 1):
                                    hdr = ws.cell(row=8, column=c).value
                                    if hdr and str(hdr).strip().startswith("D"):
                                        d_col_map[str(hdr).strip()] = c
                                
                                for cam in cams:
                                    try:
                                        m = re.search(r"-D(\d+)$", cam["cam_code"])
                                        if m:
                                            d_code = "D" + m.group(1)
                                            col_idx = d_col_map.get(d_code)
                                            if col_idx:
                                                ws.cell(row=row_idx, column=col_idx).value = 'v' if cam["status"] == 'ok' else 'F'
                                    except: pass
                                wb.save(filepath)
                            wb.close()
                        except: pass
                        break

    return RedirectResponse(url=f"/checklist/?success=1&count={checked_count}", status_code=303)


# ---------------------------------------------------------------------------
# GET /checklist/download — Download Excel files as ZIP
# ---------------------------------------------------------------------------
@router.get("/download")
@require_permission("can_export_assets")
def checklist_download(request: Request, current_user=None):
    checklist_dir = "data/camera_checklists"
    if not os.path.exists(checklist_dir):
        return RedirectResponse(url="/checklist/?error=No+Checklist+Dir", status_code=303)

    nvrs = request.query_params.getlist('nvr')

    s = io.BytesIO()
    with zipfile.ZipFile(s, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(checklist_dir):
            for file in files:
                if file.endswith('.xlsx') and not file.startswith('~'):
                    if not nvrs or any(nvr in file for nvr in nvrs):
                        filepath = os.path.join(root, file)
                        folder_name = os.path.basename(root)
                        arcname = file if folder_name == "camera_checklists" else f"{folder_name}_{file}"
                        zf.write(filepath, arcname=arcname)

    headers = {
        'Content-Disposition': 'attachment; filename="Camera_Checklists.zip"'
    }
    return StreamingResponse(
        iter([s.getvalue()]),
        media_type="application/x-zip-compressed",
        headers=headers
    )


# ===================================================================
# CRUD API Endpoints for Camera / NVR Management
# ===================================================================

@router.get("/api/cameras")
async def api_list_cameras(request: Request, db: Session = Depends(get_db)):
    """List all cameras grouped by NVR, returned as JSON."""
    current_user = get_current_user(request)
    if not current_user:
        return JSONResponse(content={"error": "Unauthorized"}, status_code=401)

    cameras = db.scalars(select(Asset).where(Asset.asset_type == "Camera").order_by(Asset.asset_name)).all()
    nvrs = db.scalars(select(Asset).where(Asset.asset_type == "NVR").order_by(Asset.asset_name)).all()

    return {
        "cameras": [
            {
                "id": c.id, "asset_code": c.asset_code, "asset_name": c.asset_name,
                "ip_address": c.ip_address, "location": c.location, "status": c.status,
                "model": c.model, "serial_number": c.serial_number, "notes": c.notes,
            }
            for c in cameras
        ],
        "nvrs": [
            {
                "id": n.id, "asset_code": n.asset_code, "asset_name": n.asset_name,
                "ip_address": n.ip_address, "location": n.location, "status": n.status,
                "model": n.model, "serial_number": n.serial_number, "notes": n.notes,
            }
            for n in nvrs
        ]
    }


@router.post("/api/camera")
async def api_create_camera(request: Request, db: Session = Depends(get_db)):
    """Create a new Camera or NVR asset."""
    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return JSONResponse(content={"error": "Unauthorized"}, status_code=403)

    data = await request.json()
    asset_type = data.get("asset_type", "Camera")  # "Camera" or "NVR"
    asset_code = data.get("asset_code", "").strip()
    asset_name = data.get("asset_name", "").strip()
    ip_address = data.get("ip_address", "").strip()
    location = data.get("location", "").strip()
    model = data.get("model", "").strip()
    serial_number = data.get("serial_number", "").strip()
    notes = data.get("notes", "").strip()

    if not asset_code or not asset_name:
        return JSONResponse(content={"error": "Mã tài sản và Tên là bắt buộc"}, status_code=400)

    # Check duplicate code
    existing = db.scalar(select(Asset).where(Asset.asset_code == asset_code))
    if existing:
        return JSONResponse(content={"error": f"Mã tài sản '{asset_code}' đã tồn tại"}, status_code=400)

    new_asset = Asset(
        asset_code=asset_code,
        asset_name=asset_name,
        asset_type=asset_type,
        ip_address=ip_address or None,
        location=location or None,
        model=model or None,
        serial_number=serial_number or None,
        notes=notes or None,
        status="active",
    )
    db.add(new_asset)
    db.commit()
    db.refresh(new_asset)

    return {
        "success": True,
        "asset": {
            "id": new_asset.id, "asset_code": new_asset.asset_code,
            "asset_name": new_asset.asset_name, "ip_address": new_asset.ip_address,
            "location": new_asset.location, "asset_type": new_asset.asset_type,
        }
    }


@router.post("/api/camera/bulk")
async def api_bulk_camera(request: Request, db: Session = Depends(get_db)):
    """Bulk manage Camera/NVR assets."""
    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return JSONResponse(content={"error": "Unauthorized"}, status_code=403)

    data = await request.json()
    asset_ids = data.get("asset_ids", [])
    action = data.get("action", "")

    if not isinstance(asset_ids, list) or not asset_ids:
        return JSONResponse(content={"error": "Không có thiết bị nào được chọn"}, status_code=400)

    assets = db.scalars(select(Asset).where(Asset.id.in_(asset_ids))).all()
    if not assets:
        return JSONResponse(content={"error": "Không tìm thấy tài sản hợp lệ"}, status_code=404)

    updated = 0
    skipped = 0

    if action == "update":
        updates = data.get("updates", {})
        for asset in assets:
            if asset.asset_type not in ["Camera", "NVR"]:
                skipped += 1
                continue
            if "location" in updates and updates["location"] is not None:
                asset.location = updates["location"].strip() or None
            if "status" in updates and updates["status"] is not None:
                asset.status = updates["status"].strip() or "active"
            updated += 1
    elif action == "activate":
        for asset in assets:
            if asset.asset_type not in ["Camera", "NVR"]:
                skipped += 1
                continue
            asset.status = "active"
            updated += 1
    elif action == "deactivate":
        for asset in assets:
            if asset.asset_type not in ["Camera", "NVR"]:
                skipped += 1
                continue
            asset.status = "retired"
            updated += 1
    else:
        return JSONResponse(content={"error": "Hành động không hợp lệ", "success": False}, status_code=400)

    db.commit()
    return {"success": True, "updated": updated, "skipped": skipped}


@router.put("/api/camera/{asset_id}")
async def api_update_camera(asset_id: int, request: Request, db: Session = Depends(get_db)):
    """Update an existing Camera or NVR asset."""
    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return JSONResponse(content={"error": "Unauthorized"}, status_code=403)

    data = await request.json()
    asset = db.get(Asset, asset_id)
    if not asset:
        return JSONResponse(content={"error": "Không tìm thấy tài sản"}, status_code=404)

    if "asset_name" in data and data["asset_name"].strip():
        asset.asset_name = data["asset_name"].strip()
    if "ip_address" in data:
        asset.ip_address = data["ip_address"].strip() or None
    if "location" in data:
        asset.location = data["location"].strip() or None
    if "model" in data:
        asset.model = data["model"].strip() or None
    if "serial_number" in data:
        asset.serial_number = data["serial_number"].strip() or None
    if "notes" in data:
        asset.notes = data["notes"].strip() or None
    if "asset_code" in data and data["asset_code"].strip():
        new_code = data["asset_code"].strip()
        if new_code != asset.asset_code:
            dup = db.scalar(select(Asset).where(Asset.asset_code == new_code))
            if dup:
                return JSONResponse(content={"error": f"Mã tài sản '{new_code}' đã tồn tại"}, status_code=400)
            asset.asset_code = new_code

    db.commit()
    return {
        "success": True,
        "asset": {
            "id": asset.id, "asset_code": asset.asset_code,
            "asset_name": asset.asset_name, "ip_address": asset.ip_address,
            "location": asset.location, "asset_type": asset.asset_type,
        }
    }


@router.delete("/api/camera/{asset_id}")
async def api_delete_camera(asset_id: int, request: Request, db: Session = Depends(get_db)):
    """Delete a Camera or NVR asset."""
    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return JSONResponse(content={"error": "Unauthorized"}, status_code=403)

    asset = db.get(Asset, asset_id)
    if not asset:
        return JSONResponse(content={"error": "Không tìm thấy tài sản"}, status_code=404)

    db.delete(asset)
    db.commit()
    return {"success": True, "deleted_id": asset_id}


# ---------------------------------------------------------------------------
# CAMERA IMPORT / EXPORT
# ---------------------------------------------------------------------------

@router.get("/import", response_class=HTMLResponse)
@require_permission("can_edit_assets")
def checklist_import_page(request: Request, current_user=None):
    return templates.TemplateResponse("assets/camera_import.html", {
        "request": request,
        "current_user": current_user,
    })

@router.post("/import")
@require_permission("can_edit_assets")
async def checklist_import_action(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), current_user=None):
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        # Standardize column names (Header is in row 1)
        headers = [str(ws.cell(row=1, column=c).value).strip() if ws.cell(row=1, column=c).value else "" for c in range(1, ws.max_column + 1)]
        
        # Headers map (Vietnamese to internal)
        header_map = {
            "Loại thiết bị": "asset_type",
            "Mã thiết bị": "asset_code",
            "Tên thiết bị": "asset_name",
            "Địa chỉ IP": "ip_address",
            "IP Đầu ghi (NVR)": "location",
            "Bộ phận": "department",
            "Model": "model",
            "Serial": "serial_number",
            "Trạng thái": "status",
            "Ghi chú": "notes"
        }
        
        # Reverse map to find column index for each internal field
        col_idx_map = {} # { "asset_code": 2, ... }
        for en_h, en_k in header_map.items():
            if en_h in headers:
                col_idx_map[en_k] = headers.index(en_h) + 1
        
        summary = {"total_rows": 0, "created": 0, "updated": 0}
        
        # Process from row 2
        for r in range(2, ws.max_row + 1):
            if not ws.cell(row=r, column=1).value and not ws.cell(row=r, column=2).value:
                continue # Skip empty rows
                
            summary["total_rows"] += 1
            
            # Extract data using map
            data = {}
            for key, idx in col_idx_map.items():
                val = ws.cell(row=r, column=idx).value
                data[key] = str(val).strip() if val is not None else None
            
            asset_code = data.get("asset_code")
            if not asset_code:
                continue
                
            # Find existing
            asset = db.scalar(select(Asset).where(Asset.asset_code == asset_code))
            if asset:
                # Update
                for k, v in data.items():
                    if k != "asset_code" and v is not None:
                        setattr(asset, k, v)
                summary["updated"] += 1
            else:
                # Create
                new_asset = Asset(
                    asset_code=asset_code,
                    asset_name=data.get("asset_name") or "Unnamed Camera",
                    asset_type=data.get("asset_type") or "Camera",
                    ip_address=data.get("ip_address"),
                    location=data.get("location"),
                    model=data.get("model"),
                    serial_number=data.get("serial_number"),
                    notes=data.get("notes"),
                    status="active"
                )
                db.add(new_asset)
                summary["created"] += 1
        
        db.commit()
        return templates.TemplateResponse("assets/camera_import.html", {
            "request": request,
            "summary": summary,
            "current_user": get_current_user(request)
        })
        
    except Exception as e:
        return templates.TemplateResponse("assets/camera_import.html", {
            "request": request,
            "error": f"Lỗi xử lý file (OpenPyXL): {str(e)}",
            "current_user": get_current_user(request)
        })

@router.get("/import/template")
@require_permission("can_edit_assets")
def checklist_import_template(request: Request, current_user=None):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Camera_NVR_Template"
    
    headers = ["Loại thiết bị", "Mã thiết bị", "Tên thiết bị", "Địa chỉ IP", "IP Đầu ghi (NVR)", "Bộ phận", "Model", "Serial", "Trạng thái", "Ghi chú"]
    ws.append(headers)
    
    # Add examples
    ws.append(["NVR", "NVR-01", "Dau ghi Cang", "192.168.1.10", "", "IT", "HIK-7608", "SN123", "active", "Main Gate"])
    ws.append(["Camera", "CAM-01-D1", "Camera Gate 1", "192.168.1.101", "192.168.1.10", "IT", "DS-2CD", "SN456", "active", ""])
    
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"camera_import_template.xlsx\""}
    )

@router.get("/export-cameras")
@require_permission("can_export_assets")
def checklist_export_cameras(request: Request, db: Session = Depends(get_db), current_user=None):
    # Export only Cameras and NVRs
    assets = db.scalars(select(Asset).where(Asset.asset_type.in_(["Camera", "NVR"])).order_by(Asset.asset_type, Asset.asset_name)).all()
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Camera_List"
    
    headers = ["Loại thiết bị", "Mã thiết bị", "Tên thiết bị", "Địa chỉ IP", "IP Đầu ghi (NVR)", "Bộ phận", "Model", "Serial", "Trạng thái", "Ghi chú"]
    ws.append(headers)
    
    for a in assets:
        ws.append([
            a.asset_type, a.asset_code, a.asset_name, a.ip_address, a.location, a.department, a.model, a.serial_number, a.status, a.notes
        ])
        
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"Camera_List_Export.xlsx\""}
    )

# ---------------------------------------------------------------------------
# CHECKLIST REPORT EXPORT
# ---------------------------------------------------------------------------

def _get_report_data_context(db: Session, start_date: str, end_date: str, nvr_ip: list[str] = None):
    """Internal helper to fetch all data needed for a report."""
    # 1. Parse dates and generate full range
    start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    
    all_dates = []
    curr = start_dt
    while curr <= end_dt:
        all_dates.append(curr.strftime("%Y-%m-%d"))
        curr += datetime.timedelta(days=1)
    
    # 2. Get cameras and group by NVR
    query = select(Asset).where(Asset.asset_type == "Camera")
    if nvr_ip:
        query = query.where(Asset.location.in_(nvr_ip))
    
    cameras = db.scalars(query.order_by(Asset.location, Asset.asset_name)).all()
    if not cameras:
        return None
        
    cam_ids = [c.id for c in cameras]
    nvr_to_cams = defaultdict(list)
    for cam in cameras:
        nvr_key = cam.location or "Chưa phân loại"
        nvr_to_cams[nvr_key].append(cam)
        
    # 3. Fetch events in range
    events = db.execute(
        select(AssetEvent)
        .where(AssetEvent.asset_id.in_(cam_ids))
        .where(AssetEvent.event_type == "daily_checklist")
        .where(AssetEvent.created_at >= start_dt)
        .where(AssetEvent.created_at <= end_dt)
        .order_by(AssetEvent.created_at.asc())
    ).scalars().all()
    
    # Chronological data pivot: Date -> { CamID -> {"status": status, "note": note} }
    report_data = defaultdict(dict)
    for ev in events:
        date_str = ev.created_at.strftime("%Y-%m-%d")
        st_text = ev.title.lower()
        if "ok" in st_text or "tốt" in st_text:
            status = "✓"
        elif "err" in st_text or "lỗi" in st_text:
            status = "F"
        else:
            status = ev.title.split(": ")[-1] if ": " in ev.title else ev.title
        
        report_data[date_str][ev.asset_id] = {
            "status": status,
            "note": ev.description or ""
        }
    
    return {
        "all_dates": all_dates,
        "nvr_to_cams": nvr_to_cams,
        "report_data": report_data,
        "start_date": start_date,
        "end_date": end_date
    }

@router.get("/export-report")
@require_permission("can_export_assets")
def checklist_export_report(
    request: Request,
    start_date: str, 
    end_date: str, 
    nvr_ip: list[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=None
):
    try:
        ctx = _get_report_data_context(db, start_date, end_date, nvr_ip)
        if not ctx:
            return RedirectResponse(url="/checklist/?error=No+cameras+found+for+selected+NVRs", status_code=303)
            
        all_dates = ctx["all_dates"]
        nvr_to_cams = ctx["nvr_to_cams"]
        report_data = ctx["report_data"]
            
        # 4. Build Excel with multiple sheets
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define Styles
        title_font = Font(size=14, bold=True, color="FFFFFF")
        header_font = Font(size=10, bold=True)
        status_font_ok = Font(size=11, bold=True, color="008000") # Green
        status_font_err = Font(size=11, bold=True, color="FF0000") # Red
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_side = Side(style='thin', color="000000")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        fill_title = PatternFill(start_color="3525CD", end_color="3525CD", fill_type="solid")
        fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for nvr_name, nvr_cameras in sorted(nvr_to_cams.items()):
            # Sanitize sheet name (31 chars max, no : \ / ? * [ ])
            safe_name = re.sub(r'[\\/*?:\[\]]', '', nvr_name)[:31]
            ws = wb.create_sheet(title=safe_name)
            
            # Header Layout
            # R1: MAIN TITLE (Merged)
            # R2: Info Headers
            # Data rows...
            
            num_cams = len(nvr_cameras)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cams + 1)
            cell_title = ws.cell(row=1, column=1)
            cell_title.value = f"BÁO CÁO CHI TIẾT CAMERA - {nvr_name} ({start_date} -> {end_date})"
            cell_title.font = title_font
            cell_title.fill = fill_title
            cell_title.alignment = center_align
            
            # Row 2: Headers
            headers = ["Ngày / Camera"] + [c.asset_name for c in nvr_cameras]
            for c_idx, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=c_idx)
                cell.value = h_text
                cell.font = header_font
                cell.fill = fill_header
                cell.border = thin_border
                cell.alignment = center_align
                
            # Row 3: IP Address sub-header
            ws.cell(row=3, column=1).value = "Địa chỉ IP"
            ws.cell(row=3, column=1).font = header_font
            ws.cell(row=3, column=1).border = thin_border
            ws.cell(row=3, column=1).fill = fill_header
            
            for c_idx, cam in enumerate(nvr_cameras, start=2):
                cell = ws.cell(row=3, column=c_idx)
                cell.value = cam.ip_address or ""
                cell.font = Font(size=9, italic=True)
                cell.border = thin_border
                cell.alignment = center_align
                
            # Data Rows (All dates)
            curr_row = 4
            for d_str in all_dates:
                # Date Cell
                date_cell = ws.cell(row=curr_row, column=1)
                date_cell.value = d_str
                date_cell.border = thin_border
                date_cell.alignment = center_align
                
                # Camera Status Cells
                for c_idx, cam in enumerate(nvr_cameras, start=2):
                    data = report_data[d_str].get(cam.id, {"status": "", "note": ""})
                    status = data["status"]
                    note = data["note"]
                    
                    cell = ws.cell(row=curr_row, column=c_idx)
                    if status and note:
                        cell.value = f"{status}\n({note})"
                    else:
                        cell.value = status
                        
                    cell.border = thin_border
                    cell.alignment = center_align
                    
                    if status == "✓":
                        cell.font = status_font_ok
                    elif status == "F":
                        cell.font = status_font_err
                
                curr_row += 1
                
            # Freeze Panes (Freeze rows 1-3 and col A)
            ws.freeze_panes = "B4"
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            for c in range(2, num_cams + 2):
                ws.column_dimensions[get_column_letter(c)].width = 14
                
        wb.save(output)
        output.seek(0)
        
        filename = f"Bao_Cao_Camera_{start_date}_{end_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
        
    except Exception as e:
        print(f"Report Export Error: {e}")
        return RedirectResponse(url=f"/checklist/?error=Report+Export+Failed: {str(e)}", status_code=303)

@router.get("/api/auto-check")
async def api_auto_check(request: Request, db: Session = Depends(get_db)):
    """Kết nối từng NVR qua Hikvision ISAPI, lấy trạng thái từng kênh camera."""
    import urllib.parse

    current_user = get_current_user(request)
    if not current_user:
        return JSONResponse(content={"error": "Unauthorized"}, status_code=401)

    try:
        nvrs = db.scalars(select(Asset).where(Asset.asset_type == "NVR")).all()
        cameras = db.scalars(select(Asset).where(Asset.asset_type == "Camera")).all()
        resources = db.scalars(select(Resource).where(Resource.is_active == True)).all()  # noqa: E712
    except Exception as exc:
        print(f"[auto-check] DB error: {exc}")
        return JSONResponse(content={"error": f"DB error: {str(exc)}"}, status_code=500)

    # camera_code → camera_id
    cam_by_code: dict[str, int] = {c.asset_code: c.id for c in cameras}

    # NVR ip_address → credentials: match by netloc or hostname (port-agnostic fallback)
    nvr_creds: dict[str, dict] = {}
    for res in resources:
        if not res.url or not res.username_hint:
            continue
        try:
            url_str = res.url.strip()
            if not url_str.startswith(('http://', 'https://')):
                url_str = 'http://' + url_str
            parsed = urllib.parse.urlparse(url_str)
            url_netloc = parsed.netloc
            url_hostname = parsed.hostname or ""
        except Exception:
            url_netloc = res.url.strip()
            url_hostname = url_netloc
        for nvr in nvrs:
            nvr_ip = (nvr.ip_address or "").strip()
            if not nvr_ip or nvr_ip in nvr_creds:
                continue
            nvr_hostname = urllib.parse.urlparse(f"http://{nvr_ip}").hostname or nvr_ip
            if url_netloc == nvr_ip or url_hostname == nvr_hostname:
                nvr_creds[nvr_ip] = {
                    "username": res.username_hint,
                    "password": decrypt_resource_password(res.password_hint) or "",
                    "url": res.url.rstrip("/"),
                }

    print(f"[auto-check] NVRs={len(nvrs)} Cameras={len(cameras)} Resources={len(resources)} CredMatches={len(nvr_creds)}")

    results: dict[str, str] = {}
    summary = {"ok": 0, "err": 0, "unknown": 0, "nvrs_checked": 0, "nvrs_failed": 0, "nvrs_no_creds": 0}

    async def _check_nvr(nvr: Asset):
        nvr_ip = (nvr.ip_address or "").strip()
        nvr_ip_code = nvr_ip.replace(".", "_").replace(":", "_")
        nvr_cameras = [c for c in cameras if (c.location or "").strip() == nvr_ip]

        if nvr_ip not in nvr_creds:
            for cam in nvr_cameras:
                results[str(cam.id)] = "unknown"
            summary["nvrs_no_creds"] += 1
            return

        creds = nvr_creds[nvr_ip]
        base_url = creds["url"] if creds["url"].startswith("http") else f"http://{creds['url']}"

        try:
            hx_timeout = httpx.Timeout(connect=8.0, read=10.0, write=5.0, pool=5.0)
            async with httpx.AsyncClient(
                auth=httpx.DigestAuth(creds["username"], creds["password"]),
                base_url=base_url, verify=False, timeout=hx_timeout
            ) as client:
                xml_texts = []
                for ep in [
                    "/ISAPI/ContentMgmt/InputProxy/channels/status",
                    "/ISAPI/System/Video/inputs/channels/status",
                ]:
                    try:
                        resp = await client.get(ep)
                        if resp.status_code == 200:
                            xml_texts.append(resp.text)
                    except Exception:
                        continue

                if not xml_texts:
                    for cam in nvr_cameras:
                        results[str(cam.id)] = "unknown"
                    summary["nvrs_failed"] += 1
                    print(f"[auto-check] NVR {nvr_ip}: no XML from either endpoint")
                    return

                items = []
                for xml_text in xml_texts:
                    try:
                        clean_xml = re.sub(r' xmlns="[^"]+"', '', xml_text)
                        root = ET.fromstring(clean_xml)
                        items.extend(
                            root.findall(".//InputProxyChannelStatus")
                            + root.findall(".//VideoInputStatus")
                            + root.findall(".//VideoInputChannelStatus")
                            + root.findall(".//VideoInputChannel")
                            + root.findall(".//InputProxyChannel")
                        )
                    except Exception as parse_e:
                        print(f"[auto-check] error parsing XML: {parse_e}")
                        continue

                for item in items:
                    ch_id = item.findtext("id") or item.findtext("channelID")
                    if not ch_id:
                        continue
                        
                    online = (item.findtext("online") or item.findtext("videoInputEnabled") or "").strip().lower()
                    signalState = (item.findtext("signalStatus") or item.findtext("videoSignalStatus") or item.findtext("resDesc") or "").strip().lower()
                    videoLoss = (item.findtext("videoLoss") or "").strip().lower()

                    is_ok = True
                    if online == "false":
                        is_ok = False
                    if signalState in ("loss", "disconnected", "false", "no video") or "loss" in signalState or "disconnect" in signalState:
                        is_ok = False
                    if videoLoss == "true":
                        is_ok = False
                    cam_code = f"CAM-{nvr_ip_code}-D{ch_id}"
                    cam_id = cam_by_code.get(cam_code)
                    if cam_id:
                        results[str(cam_id)] = "ok" if is_ok else "err"

                for cam in nvr_cameras:
                    if str(cam.id) not in results:
                        results[str(cam.id)] = "unknown"

                summary["nvrs_checked"] += 1
                print(f"[auto-check] NVR {nvr_ip}: OK, {len(items)} channels parsed")

        except Exception as exc:
            for cam in nvr_cameras:
                results[str(cam.id)] = "unknown"
            summary["nvrs_failed"] += 1
            print(f"[auto-check] NVR {nvr_ip} error: {exc}")

    async def _check_nvr_guarded(nvr: Asset):
        """Hard 15s deadline per NVR — prevents hang if httpx stalls."""
        nvr_ip = (nvr.ip_address or "").strip()
        try:
            await asyncio.wait_for(_check_nvr(nvr), timeout=15.0)
        except asyncio.TimeoutError:
            nvr_cameras = [c for c in cameras if (c.location or "").strip() == nvr_ip]
            for cam in nvr_cameras:
                if str(cam.id) not in results:
                    results[str(cam.id)] = "unknown"
            summary["nvrs_failed"] += 1
            print(f"[auto-check] NVR {nvr_ip}: hard timeout (15s)")

    try:
        await asyncio.gather(*[_check_nvr_guarded(nvr) for nvr in nvrs])
    except Exception as exc:
        print(f"[auto-check] gather error: {exc}")

    for v in results.values():
        if v in summary:
            summary[v] += 1

    print(f"[auto-check] Done: {summary}")
    return JSONResponse(content={"results": results, "summary": summary})


@router.post("/api/sync-cameras")
async def api_sync_cameras(request: Request, db: Session = Depends(get_db)):
    """Quét từng NVR qua ISAPI, đồng bộ danh sách camera vào DB."""
    import urllib.parse

    current_user = get_current_user(request)
    if not current_user or not has_permission(current_user, "can_edit_assets"):
        return JSONResponse(content={"error": "Unauthorized"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        body = {}
    overwrite_names: bool = body.get("overwrite_names", True)
    dry_run: bool = body.get("dry_run", False)

    try:
        nvrs = db.scalars(select(Asset).where(Asset.asset_type == "NVR")).all()
        cameras = db.scalars(select(Asset).where(Asset.asset_type == "Camera")).all()
        resources = db.scalars(select(Resource).where(Resource.is_active == True)).all()  # noqa: E712
    except Exception as exc:
        print(f"[sync] DB error: {exc}")
        return JSONResponse(content={"error": f"DB error: {str(exc)}"}, status_code=500)

    # Build credential map: match Resource URL against NVR ip_address
    # Tries netloc match first (exact host:port), then hostname-only fallback
    nvr_creds: dict[str, dict] = {}
    for res in resources:
        if not res.url or not res.username_hint:
            continue
        try:
            url_str = res.url.strip()
            if not url_str.startswith(("http://", "https://")):
                url_str = "http://" + url_str
            parsed = urllib.parse.urlparse(url_str)
            url_netloc = parsed.netloc          # "host:port" or "host"
            url_hostname = parsed.hostname or ""  # "host" only (no port)
        except Exception:
            url_netloc = res.url.strip()
            url_hostname = url_netloc
        for nvr in nvrs:
            nvr_ip = (nvr.ip_address or "").strip()
            if not nvr_ip or nvr_ip in nvr_creds:
                continue
            # Match by netloc (exact), or by hostname only (ignoring port difference)
            nvr_hostname = urllib.parse.urlparse(f"http://{nvr_ip}").hostname or nvr_ip
            if url_netloc == nvr_ip or url_hostname == nvr_hostname:
                nvr_creds[nvr_ip] = {
                    "username": res.username_hint,
                    "password": decrypt_resource_password(res.password_hint) or "",
                    "url": res.url.rstrip("/"),
                }
                print(f"[sync] matched NVR {nvr_ip!r} via resource URL {res.url!r}")

    # asset_code → Asset object (mutable for new inserts)
    cam_by_code: dict[str, Asset] = {c.asset_code: c for c in cameras}

    summary = {
        "nvrs_synced": 0, "nvrs_failed": 0, "nvrs_no_creds": 0,
        "total_created": 0, "total_updated": 0, "total_deactivated": 0,
        "dry_run": dry_run,
    }
    details: list[dict] = []

    async def _sync_nvr(nvr: Asset):
        nvr_ip = (nvr.ip_address or "").strip()
        nvr_ip_code = nvr_ip.replace(".", "_").replace(":", "_")
        nvr_label = nvr.asset_name or nvr_ip

        if nvr_ip not in nvr_creds:
            summary["nvrs_no_creds"] += 1
            details.append({"nvr": nvr_label, "ip": nvr_ip, "status": "no_creds",
                            "created": 0, "updated": 0, "deactivated": 0})
            return

        creds = nvr_creds[nvr_ip]
        base_url = creds["url"] if creds["url"].startswith("http") else f"http://{creds['url']}"

        try:
            hx_timeout = httpx.Timeout(connect=8.0, read=12.0, write=5.0, pool=5.0)
            async with httpx.AsyncClient(
                auth=httpx.DigestAuth(creds["username"], creds["password"]),
                base_url=base_url, verify=False, timeout=hx_timeout
            ) as client:
                xml_text = None
                for ep in [
                    "/ISAPI/ContentMgmt/InputProxy/channels",
                    "/ISAPI/System/Video/inputs/channels",
                ]:
                    try:
                        resp = await client.get(ep)
                        if resp.status_code == 200:
                            xml_text = resp.text
                            break
                    except Exception:
                        continue

                if not xml_text:
                    summary["nvrs_failed"] += 1
                    details.append({"nvr": nvr_label, "ip": nvr_ip, "status": "failed",
                                    "created": 0, "updated": 0, "deactivated": 0})
                    print(f"[sync] NVR {nvr_ip}: no XML from either channel-list endpoint")
                    return

                clean_xml = re.sub(r' xmlns="[^"]+"', "", xml_text)
                root = ET.fromstring(clean_xml)
                channel_items = (
                    root.findall(".//InputProxyChannel")
                    or root.findall(".//VideoInputChannel")
                )

                if not channel_items:
                    summary["nvrs_failed"] += 1
                    details.append({"nvr": nvr_label, "ip": nvr_ip, "status": "no_channels",
                                    "created": 0, "updated": 0, "deactivated": 0})
                    print(f"[sync] NVR {nvr_ip}: XML OK but no channel items found")
                    return

                nvr_channel_codes: set[str] = set()
                created = updated = deactivated = 0

                for item in channel_items:
                    ch_id = item.findtext("id") or item.findtext("channelID")
                    ch_name = (item.findtext("name") or "").strip()
                    if not ch_id:
                        continue

                    cam_code = f"CAM-{nvr_ip_code}-D{ch_id}"
                    nvr_channel_codes.add(cam_code)

                    existing = cam_by_code.get(cam_code)
                    if not dry_run:
                        if existing:
                            if overwrite_names and ch_name:
                                existing.asset_name = ch_name
                            existing.location = nvr_ip
                            existing.status = "active"
                            updated += 1
                        else:
                            new_cam = Asset(
                                asset_code=cam_code,
                                asset_name=ch_name or cam_code,
                                asset_type="Camera",
                                location=nvr_ip,
                                status="active",
                            )
                            db.add(new_cam)
                            cam_by_code[cam_code] = new_cam
                            created += 1
                    else:
                        if existing:
                            updated += 1
                        else:
                            created += 1

                # Cameras in DB for this NVR but absent from NVR response → inactive
                for cam in cameras:
                    if (cam.location or "").strip() == nvr_ip and cam.asset_code not in nvr_channel_codes:
                        if not dry_run:
                            cam.status = "inactive"
                        deactivated += 1

                summary["nvrs_synced"] += 1
                summary["total_created"] += created
                summary["total_updated"] += updated
                summary["total_deactivated"] += deactivated
                details.append({
                    "nvr": nvr_label, "ip": nvr_ip, "status": "ok",
                    "channels": len(channel_items),
                    "created": created, "updated": updated, "deactivated": deactivated,
                })
                print(f"[sync] NVR {nvr_ip}: +{created} created, ~{updated} updated, -{deactivated} deactivated")

        except Exception as exc:
            summary["nvrs_failed"] += 1
            details.append({"nvr": nvr_label, "ip": nvr_ip, "status": "error",
                            "error": str(exc), "created": 0, "updated": 0, "deactivated": 0})
            print(f"[sync] NVR {nvr_ip} error: {exc}")

    async def _sync_nvr_guarded(nvr: Asset):
        nvr_ip = (nvr.ip_address or "").strip()
        try:
            await asyncio.wait_for(_sync_nvr(nvr), timeout=20.0)
        except asyncio.TimeoutError:
            summary["nvrs_failed"] += 1
            details.append({"nvr": nvr.asset_name or nvr_ip, "ip": nvr_ip, "status": "timeout",
                            "created": 0, "updated": 0, "deactivated": 0})
            print(f"[sync] NVR {nvr_ip}: hard timeout (20s)")

    await asyncio.gather(*[_sync_nvr_guarded(nvr) for nvr in nvrs])

    if not dry_run:
        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            print(f"[sync] commit error: {exc}")
            return JSONResponse(content={"error": f"DB commit error: {str(exc)}"}, status_code=500)

    print(f"[sync] Done: {summary}")
    return JSONResponse(content={"summary": summary, "details": details})


@router.get("/report/preview")
@require_permission("can_export_assets")
def checklist_report_preview(
    request: Request,
    start_date: str, 
    end_date: str, 
    nvr_ip: list[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=None
):
    try:
        ctx = _get_report_data_context(db, start_date, end_date, nvr_ip)
        if not ctx:
            return HTMLResponse(content="<div class='p-8 text-center text-on-surface/50 font-medium'>Không tìm thấy dữ liệu Camera cho các đầu ghi đã chọn.</div>")
            
        return templates.TemplateResponse("assets/report_preview_table.html", {
            "request": request,
            **ctx
        })
    except Exception as e:
        return HTMLResponse(content=f"<div class='p-8 text-rose-500 font-bold'>Lỗi: {str(e)}</div>")
