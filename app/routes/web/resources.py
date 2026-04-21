import io
import json
import math
import re
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
import openpyxl
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.auth import require_module_access, require_permission
from app.db.models import Resource
from app.db.session import get_db
from app.services.audit import log_audit
from app.security import encrypt_resource_password, decrypt_resource_password

router = APIRouter(prefix='/resources', tags=['resources'])
templates = Jinja2Templates(directory='app/templates')
PAGE_SIZE = 10


def _render_resource_form(request: Request, current_user, *, item=None, error: str | None = None):
    return templates.TemplateResponse('resources/form.html', {'request': request, 'item': item, 'error': error, 'current_user': current_user})


def _render_import_page(request: Request, current_user, *, error: str | None = None, summary: dict | None = None):
    return templates.TemplateResponse('resources/import.html', {'request': request, 'current_user': current_user, 'error': error, 'summary': summary})


def _client_ip(request: Request) -> str:
    forwarded_for = request.headers.get('x-forwarded-for', '').split(',')[0].strip()
    if forwarded_for:
        return forwarded_for
    return request.client.host if request.client else 'unknown'


def _normalize_ip(value) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if '.' in raw:
        return raw
    digits = ''.join(ch for ch in raw if ch.isdigit())
    if len(digits) == 12:
        parts = [digits[i:i + 3] for i in range(0, 12, 3)]
        return '.'.join(str(int(part)) for part in parts)
    return raw


def _parse_user_password_block(value) -> tuple[str | None, str | None]:
    raw = str(value or '').strip()
    if not raw:
        return None, None
    user_match = re.search(r'user\s*:\s*(.+)', raw, flags=re.IGNORECASE)
    pass_match = re.search(r'password\s*:\s*(.+)', raw, flags=re.IGNORECASE)
    username = user_match.group(1).strip() if user_match else None
    password = pass_match.group(1).strip() if pass_match else None
    return username or None, password or None


def _parse_resource_export_excel(rows: list[tuple]) -> list[dict]:
    headers = [str(value).strip() if value is not None else '' for value in rows[0]]
    normalized = [header.lower() for header in headers]
    required = ['tên', 'nhóm', 'loại', 'url', 'user', 'password', 'ghi chú', 'file path']
    if not all(name in normalized for name in required):
        raise ValueError('File không đúng định dạng export tài nguyên.')

    index_map = {header.lower(): idx for idx, header in enumerate(headers)}
    items = []
    for row in rows[1:]:
        title = str(row[index_map['tên']]).strip() if len(row) > index_map['tên'] and row[index_map['tên']] is not None else ''
        if not title:
            continue
        items.append({
            'title': title,
            'category': str(row[index_map['nhóm']]).strip() if len(row) > index_map['nhóm'] and row[index_map['nhóm']] is not None else 'general',
            'resource_type': str(row[index_map['loại']]).strip() if len(row) > index_map['loại'] and row[index_map['loại']] is not None else 'web',
            'url': str(row[index_map['url']]).strip() if len(row) > index_map['url'] and row[index_map['url']] is not None else None,
            'username_hint': str(row[index_map['user']]).strip() if len(row) > index_map['user'] and row[index_map['user']] is not None else None,
            'password_hint': str(row[index_map['password']]).strip() if len(row) > index_map['password'] and row[index_map['password']] is not None else None,
            'note': str(row[index_map['ghi chú']]).strip() if len(row) > index_map['ghi chú'] and row[index_map['ghi chú']] is not None else None,
            'file_path': str(row[index_map['file path']]).strip() if len(row) > index_map['file path'] and row[index_map['file path']] is not None else None,
        })
    if not items:
        raise ValueError('File export tài nguyên không có dòng dữ liệu hợp lệ.')
    return items



def _parse_resources_source_excel(rows: list[tuple]) -> list[dict]:
    if len(rows) < 2:
        raise ValueError('File import không đủ dữ liệu.')

    current_group = None
    items = []
    for row in rows[2:]:
        first = row[0] if len(row) > 0 else None
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        ip_value = row[2] if len(row) > 2 else None
        line_code = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        credentials = row[4] if len(row) > 4 else None

        first_text = str(first).strip() if first is not None else ''
        if first_text and not name and not ip_value and not line_code and not credentials:
            current_group = first_text
            continue
        if not name:
            continue

        username, password = _parse_user_password_block(credentials)
        ip_text = _normalize_ip(ip_value)
        note_parts = []
        if current_group:
            note_parts.append(f'Khu vực: {current_group}')
        if line_code:
            note_parts.append(f'Ký hiệu dây: {line_code}')
        if ip_text:
            note_parts.append(f'IP: {ip_text}')
        note = ' | '.join(note_parts) if note_parts else None
        title = f'{current_group} - {name}' if current_group else name
        url = f'http://{ip_text}' if ip_text else None
        items.append({
            'title': title,
            'category': 'camera',
            'resource_type': 'device',
            'url': url,
            'username_hint': username,
            'password_hint': password,
            'note': note,
            'file_path': None,
            'ip_text': ip_text,
        })
    if not items:
        raise ValueError('Không tìm thấy dòng dữ liệu hợp lệ để import.')
    return items



def _parse_resources_excel(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('File import đang rỗng.')

    try:
        return _parse_resource_export_excel(rows)
    except ValueError:
        return _parse_resources_source_excel(rows)


def _import_resource_rows(db: Session, rows: list[dict]):
    created = 0
    updated = 0
    for row in rows:
        existing = db.scalar(select(Resource).where(Resource.title == row['title']))
        if existing:
            existing.category = row['category']
            existing.resource_type = row['resource_type']
            existing.url = row['url']
            existing.username_hint = row['username_hint']
            existing.password_hint = encrypt_resource_password(row['password_hint'])
            existing.note = row['note']
            existing.is_active = True
            updated += 1
        else:
            db.add(Resource(
                title=row['title'],
                category=row['category'],
                resource_type=row['resource_type'],
                url=row['url'],
                username_hint=row['username_hint'],
                password_hint=encrypt_resource_password(row['password_hint']),
                note=row['note'],
                file_path=None,
                is_active=True,
                is_hidden=True,
            ))
            created += 1
    db.commit()
    return {'created': created, 'updated': updated, 'total': len(rows)}



def _resource_base_stmt(q: str | None = None, category: str | None = None):
    stmt = select(Resource).where(Resource.is_active == True)  # noqa: E712
    if q:
        like = f'%{q.strip()}%'
        stmt = stmt.where(or_(Resource.title.ilike(like), Resource.url.ilike(like), Resource.note.ilike(like), Resource.username_hint.ilike(like)))
    if category:
        stmt = stmt.where(Resource.category == category)
    return stmt.order_by(Resource.category.asc(), Resource.title.asc())


def _decrypt_page(db: Session, items):
    dirty = False
    for item in items:
        if item.password_hint and not item.password_hint.startswith('aes:'):
            item.password_hint = encrypt_resource_password(item.password_hint)
            dirty = True
        item._decrypted_password = decrypt_resource_password(item.password_hint)
    if dirty:
        db.commit()
    return items


def _filtered_resources(db: Session, q: str | None = None, category: str | None = None):
    """Load all matching resources — dùng cho export."""
    items = db.scalars(_resource_base_stmt(q, category)).all()
    return _decrypt_page(db, items)


@router.get('/', response_class=HTMLResponse)
@require_module_access('resources')
def resource_list(request: Request, q: str | None = Query(default=None), category: str | None = Query(default=None), page: int = Query(default=1), success: str | None = Query(default=None), skipped: str | None = Query(default=None), db: Session = Depends(get_db), current_user=None):
    base = _resource_base_stmt(q, category)
    total_items = db.scalar(select(func.count()).select_from(base.subquery())) or 0
    total_pages = max(1, math.ceil(total_items / PAGE_SIZE))
    page = max(1, min(page, total_pages))
    items = db.scalars(base.limit(PAGE_SIZE).offset((page - 1) * PAGE_SIZE)).all()
    _decrypt_page(db, items)
    categories = db.scalars(select(Resource.category).where(Resource.category.is_not(None)).distinct().order_by(Resource.category.asc())).all()
    return templates.TemplateResponse('resources/list.html', {'request': request, 'items': items, 'categories': categories, 'q': q or '', 'category': category or '', 'current_user': current_user, 'page': page, 'total_pages': total_pages, 'total_items': total_items, 'success': success, 'skipped': skipped})


@router.get('/partial', response_class=HTMLResponse)
@require_module_access('resources')
def resource_list_partial(request: Request, q: str | None = Query(default=None), category: str | None = Query(default=None), page: int = Query(default=1), db: Session = Depends(get_db), current_user=None):
    base = _resource_base_stmt(q, category)
    total_items = db.scalar(select(func.count()).select_from(base.subquery())) or 0
    total_pages = max(1, math.ceil(total_items / PAGE_SIZE))
    page = max(1, min(page, total_pages))
    items = db.scalars(base.limit(PAGE_SIZE).offset((page - 1) * PAGE_SIZE)).all()
    _decrypt_page(db, items)
    return templates.TemplateResponse('resources/_table.html', {
        'request': request, 'items': items, 'q': q or '', 'category': category or '',
        'current_user': current_user, 'page': page, 'total_pages': total_pages,
    })


@router.get('/export')
@require_permission('can_manage_resources')
def resource_export(request: Request, q: str | None = Query(default=None), category: str | None = Query(default=None), db: Session = Depends(get_db), current_user=None):
    log_audit(db, actor=current_user.username if current_user else None, module='resources', action='export', entity_type='resource', entity_id=None, metadata={'q': q or None, 'category': category or None, 'ip': _client_ip(request)})
    db.commit()
    items = _filtered_resources(db, q=q, category=category)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resources'
    ws.append(['ID', 'Tên', 'Nhóm', 'Loại', 'URL', 'User', 'Password', 'Ghi chú', 'File path'])
    for item in items:
        ws.append([
            item.id,
            item.title,
            item.category or '',
            item.resource_type or '',
            item.url or '',
            item.username_hint or '',
            decrypt_resource_password(item.password_hint) or '',
            item.note or '',
            item.file_path or '',
        ])
    stream = io.BytesIO()
    wb.save(stream)
    content = stream.getvalue()
    return Response(
        content=content,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="resources_export.xlsx"',
            'Content-Length': str(len(content))
        }
    )


@router.get('/new', response_class=HTMLResponse)
@require_permission('can_manage_resources')
def resource_new(request: Request, current_user=None):
    return _render_resource_form(request, current_user, item=None, error=None)


@router.post('/new')
@require_permission('can_manage_resources')
def resource_create(request: Request, title: str = Form(...), category: str = Form(default='general'), resource_type: str = Form(default='web'), url: str = Form(default=''), username_hint: str = Form(default=''), password_hint: str = Form(default=''), note: str = Form(default=''), file_path: str = Form(default=''), db: Session = Depends(get_db), current_user=None):
    item = Resource(title=title.strip(), category=category.strip() or 'general', resource_type=resource_type.strip() or 'web', url=url.strip() or None, username_hint=username_hint.strip() or None, password_hint=encrypt_resource_password(password_hint.strip() or None), note=note.strip() or None, file_path=file_path.strip() or None, is_active=True, is_hidden=True)
    db.add(item)
    db.commit()
    return RedirectResponse('/resources/', status_code=303)


@router.get('/import', response_class=HTMLResponse)
@require_permission('can_manage_resources')
def resource_import_page(request: Request, current_user=None):
    return _render_import_page(request, current_user)


@router.post('/import', response_class=HTMLResponse)
@require_permission('can_manage_resources')
def resource_import_submit(request: Request, import_file: UploadFile = File(...), db: Session = Depends(get_db), current_user=None):
    filename = (import_file.filename or '').lower()
    if not filename.endswith('.xlsx'):
        return _render_import_page(request, current_user, error='Chỉ hỗ trợ file .xlsx')
    try:
        rows = _parse_resources_excel(import_file.file.read())
        summary = _import_resource_rows(db, rows)
    except ValueError as exc:
        return _render_import_page(request, current_user, error=str(exc))
    except Exception:
        return _render_import_page(request, current_user, error='Không đọc được file import tài nguyên. Hãy dùng đúng file Excel đã chuẩn hóa.')
    return _render_import_page(request, current_user, summary=summary)


@router.get('/{resource_id}/edit', response_class=HTMLResponse)
@require_permission('can_manage_resources')
def resource_edit(resource_id: int, request: Request, db: Session = Depends(get_db), current_user=None):
    item = db.get(Resource, resource_id)
    if item:
        item._decrypted_password = decrypt_resource_password(item.password_hint)
    return _render_resource_form(request, current_user, item=item, error=None)


@router.post('/{resource_id}/edit')
@require_permission('can_manage_resources')
def resource_update(resource_id: int, request: Request, title: str = Form(...), category: str = Form(default='general'), resource_type: str = Form(default='web'), url: str = Form(default=''), username_hint: str = Form(default=''), password_hint: str = Form(default=''), note: str = Form(default=''), file_path: str = Form(default=''), is_active: str | None = Form(default=None), db: Session = Depends(get_db), current_user=None):
    item = db.get(Resource, resource_id)
    item.title = title.strip()
    item.category = category.strip() or 'general'
    item.resource_type = resource_type.strip() or 'web'
    item.url = url.strip() or None
    item.username_hint = username_hint.strip() or None
    item.password_hint = encrypt_resource_password(password_hint.strip() or None)
    item.note = note.strip() or None
    item.file_path = file_path.strip() or None
    item.is_active = is_active == 'true'
    db.commit()
    return RedirectResponse('/resources/', status_code=303)


@router.get('/{resource_id}/password')
@require_permission('can_manage_resources')
def resource_password(resource_id: int, request: Request, db: Session = Depends(get_db), current_user=None):
    item = db.get(Resource, resource_id)
    if not item or not item.is_active:
        return Response(status_code=404)
    password_value = decrypt_resource_password(item.password_hint)
    log_audit(db, actor=current_user.username if current_user else None, module='resources', action='view_password', entity_type='resource', entity_id=item.id, metadata={'title': item.title, 'ip': _client_ip(request)})
    db.commit()
    return Response(content=json.dumps({'password': password_value or ''}, ensure_ascii=False), media_type='application/json')


@router.post('/{resource_id}/archive')
@require_permission('can_manage_resources')
def resource_archive(resource_id: int, request: Request, db: Session = Depends(get_db), current_user=None):
    item = db.get(Resource, resource_id)
    if item:
        item.is_active = False
        db.commit()
    return RedirectResponse('/resources/', status_code=303)



@router.post('/bulk-update')
@require_permission('can_manage_resources')
def resources_bulk_update(
    request: Request,
    resource_ids: str = Form(...),
    category: str | None = Form(default=None),
    is_active: str | None = Form(default=None),
    db: Session = Depends(get_db),
    current_user=None
):
    ids_list = [int(i.strip()) for i in resource_ids.split(",") if i.strip().isdigit()]
    updated = 0
    if ids_list:
        items = db.scalars(select(Resource).where(Resource.id.in_(ids_list))).all()
        for item in items:
            changed = False
            if category:
                item.category = category.strip()
                changed = True
            if is_active is not None:
                item.is_active = (is_active == 'true')
                changed = True
            if changed:
                updated += 1
                log_audit(db, actor=current_user.username if current_user else None, module='resources', action='bulk_update', entity_type='resource', entity_id=item.id, metadata={'category': category, 'is_active': is_active})
        db.commit()
    return RedirectResponse(f'/resources/?success={updated}', status_code=303)

@router.post('/bulk-archive')
@require_permission('can_manage_resources')
def resources_bulk_archive(
    request: Request,
    resource_ids: str = Form(...),
    db: Session = Depends(get_db),
    current_user=None
):
    ids_list = [int(i.strip()) for i in resource_ids.split(",") if i.strip().isdigit()]
    updated = 0
    if ids_list:
        items = db.scalars(select(Resource).where(Resource.id.in_(ids_list))).all()
        for item in items:
            item.is_active = False
            updated += 1
            log_audit(db, actor=current_user.username if current_user else None, module='resources', action='bulk_archive', entity_type='resource', entity_id=item.id)
        db.commit()
    return RedirectResponse(f'/resources/?success={updated}', status_code=303)

@router.post('/bulk-restore')
@require_permission('can_manage_resources')
def resources_bulk_restore(
    request: Request,
    resource_ids: str = Form(...),
    db: Session = Depends(get_db),
    current_user=None
):
    ids_list = [int(i.strip()) for i in resource_ids.split(",") if i.strip().isdigit()]
    updated = 0
    if ids_list:
        items = db.scalars(select(Resource).where(Resource.id.in_(ids_list))).all()
        for item in items:
            item.is_active = True
            updated += 1
            log_audit(db, actor=current_user.username if current_user else None, module='resources', action='bulk_restore', entity_type='resource', entity_id=item.id)
        db.commit()
    return RedirectResponse(f'/resources/?success={updated}', status_code=303)

@router.post('/bulk-delete-hard')
@require_permission('can_manage_resources')
def resources_bulk_delete_hard(
    request: Request,
    resource_ids: str = Form(...),
    confirm_text: str = Form(default=""),
    db: Session = Depends(get_db),
    current_user=None
):
    if not current_user or current_user.role != 'admin':
        return RedirectResponse('/resources/?error=unauthorized_admin_only', status_code=303)

    ids_list = [int(i.strip()) for i in resource_ids.split(",") if i.strip().isdigit()]
    
    if confirm_text != f"DELETE {len(ids_list)}":
        return RedirectResponse('/resources/?error=invalid_confirm', status_code=303)

    if not ids_list:
        return RedirectResponse('/resources/', status_code=303)

    items = db.scalars(select(Resource).where(Resource.id.in_(ids_list))).all()
    success_count = 0
    skipped_count = 0
    
    for item in items:
        # Check referential constraints
        # Specifically, Resource has no direct ForeignKeys attached in models currently.
        # However, checking generically.
        has_dependencies = False
        
        if has_dependencies:
            skipped_count += 1
            log_audit(db, actor=current_user.username, module='resources', action='bulk_hard_delete', entity_type='resource', entity_id=item.id, result='skipped', reason='has_dependencies')
            continue

        log_audit(db, actor=current_user.username, module='resources', action='bulk_hard_delete', entity_type='resource', entity_id=item.id, result='success')
        db.delete(item)
        success_count += 1
        
    db.commit()
    return RedirectResponse(f'/resources/?success={success_count}&skipped={skipped_count}', status_code=303)


def import_desktop_resources(db: Session):
    desktop = Path.home() / 'Desktop'
    imported = 0
    json_file = desktop / 'dashboard_backup.json'
    if json_file.exists():
        try:
            data = json.loads(json_file.read_text(encoding='utf-8'))
            if isinstance(data, list):
                for row in data:
                    title = (row.get('name') or '').strip()
                    url = (row.get('url') or '').strip()
                    if not title:
                        continue
                    exists = db.scalar(select(Resource).where(Resource.title == title, Resource.url == (url or None)))
                    if exists:
                        continue
                    category = 'camera' if 'camera' in title.lower() else 'dashboard'
                    db.add(Resource(title=title, category=category, resource_type='web', url=url or None, username_hint=(row.get('username') or '').strip() or None, password_hint=(row.get('password') or '').strip() or None, note=(row.get('note') or '').strip() or None, is_active=True, is_hidden=True))
                    imported += 1
        except Exception:
            pass
    camera_file = desktop / 'CAMERA.html'
    if camera_file.exists():
        title = 'CAMERA.html'
        exists = db.scalar(select(Resource).where(Resource.title == title, Resource.file_path == str(camera_file)))
        if not exists:
            db.add(Resource(title=title, category='camera', resource_type='file', file_path=str(camera_file), note='File desktop dùng cho support camera', is_active=True, is_hidden=True))
            imported += 1
    db.commit()
    return imported
